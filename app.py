# ==================================================== 
# IMPORTS
# ==================================================== 
from ui_components import apply_global_styling, render_kpi_card, render_enhanced_dataframe, render_empty_state, render_breadcrumb, render_page_header, COLORS
import streamlit as st
import pandas as pd
from datetime import datetime, timedelta, date
import gspread
from rapidfuzz import fuzz
from oauth2client.service_account import ServiceAccountCredentials
import os
import json
from login_master_with_branding import main as login_main
from status_updater import sync_all_statuses
from filter_candidates import render_filter_section as render_candidate_filter
from filter_companies import render_filter_section as render_company_filter
from export_utils import export_single_match, export_to_interview_sheet
from candidate_wizard_module import render_wizard
from job_matcher_module import run_matching, export_to_interview_sheet
import warnings
warnings.filterwarnings('ignore')
import secrets
import string
import hashlib
# from agency_managment import admin_agency_management, agency_user_management, get_all_agencies
from openpyxl.utils import get_column_letter

if "page" not in st.session_state:
    st.session_state.page = "landing"

current_page = st.session_state.get("page", "landing")


# ==================================================== 
# PAGE CONFIG
# ==================================================== 
st.set_page_config(
    page_title="Placify - Placement Agency",
    page_icon="",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
header[data-testid="stHeader"] {
    height: 0px !important;
    min-height: 0px !important;
    padding: 0 !important;
}

.main .block-container {
    max-width: 1200px;
    padding-top: 1.5rem;
    padding-bottom: 2rem;
    margin-left: auto;
    margin-right: auto;
}
</style>
""", unsafe_allow_html=True)
# ============ CONDITIONAL CSS ============
# if current_page in ["landing", "accept_tc"]:
#     st.markdown("""
#       <style>
#         [data-testid="stSidebarNav"] { display: none !important; }
#         [data-testid="stSidebar"] { display: none !important; }
#         .main .block-container { max-width: 800px !important; margin: 0 auto !important; }
#         [data-testid="stAppViewContainer"] { padding-left: 2rem !important; padding-right: 2rem !important; }
#       </style>
#       """, unsafe_allow_html=True)


# ==================================================== 
# SIDEBAR STYLING & RENDERING
# ==================================================== 

def apply_custom_sidebar_css():
    """Apply modern sidebar styling - clean white with indigo accents"""
    css = """
    <style>
    /* Sidebar container */
    [data-testid="stSidebar"] {
        background: #FFFFFF !important;
        border-right: 1px solid #E2E8F0 !important;
    }
    [data-testid="stSidebar"] [data-testid="stVerticalBlockBG"] {
        background: transparent !important;
    }
    [data-testid="stSidebar"] * {
        color: #334155 !important;
    }

    /* Hide radio circle indicators */
    [data-testid="stSidebar"] .stRadio input[type="radio"] {
        display: none !important;
    }

    /* Nav items */
    [data-testid="stSidebar"] .stRadio label {
        color: #64748B !important;
        font-weight: 500 !important;
        padding: 10px 16px !important;
        margin: 2px 8px !important;
        border-radius: 8px !important;
        cursor: pointer !important;
        background-color: transparent !important;
        border: none !important;
        display: flex !important;
        align-items: center !important;
        font-size: 14px !important;
        transition: all 0.15s ease !important;
        min-height: 40px !important;
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif !important;
    }
    [data-testid="stSidebar"] .stRadio label:hover {
        background-color: #F1F5F9 !important;
        color: #0F172A !important;
    }
    [data-testid="stSidebar"] .stRadio input[type="radio"]:checked + label,
    [data-testid="stSidebar"] .stRadio label[data-checked="true"] {
        background-color: #EEF2FF !important;
        color: #6366F1 !important;
        font-weight: 600 !important;
        border-left: 3px solid #6366F1 !important;
    }

    /* Sidebar radio group label */
    [data-testid="stSidebar"] .stRadio > label {
        display: none !important;
    }

    /* Sidebar buttons (logout) */
    [data-testid="stSidebar"] .stButton > button {
        background: #F8FAFC !important;
        color: #64748B !important;
        border: 1px solid #E2E8F0 !important;
        border-radius: 8px !important;
        padding: 10px 20px !important;
        font-weight: 600 !important;
        width: 100% !important;
        font-size: 13px !important;
        letter-spacing: 0.02em !important;
        transition: all 0.15s ease !important;
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif !important;
    }
    [data-testid="stSidebar"] .stButton > button:hover {
        background: #FEF2F2 !important;
        color: #EF4444 !important;
        border-color: #EF4444 !important;
        transform: none !important;
    }

    /* Sidebar divider */
    [data-testid="stSidebar"] hr {
        border: none !important;
        height: 1px !important;
        background: #E2E8F0 !important;
        margin: 10px 16px !important;
    }
    </style>
    """
    st.markdown(css, unsafe_allow_html=True)

# ====================================================
# LOGOUT FUNCTION - ADD TO APP.PY
# ====================================================
# Paste this function near the top of app.py (after imports)
# Around line 50-100

# def logout():
#     """
#     Logout function - clears all session state
#     """
#     # Get list of all session state keys
#     keys_to_delete = list(st.session_state.keys())
    
#     # Delete each key
#     for key in keys_to_delete:
#         try:
#             del st.session_state[key]
#         except:
#             pass
    
#     # Rerun to show login page
#     st.rerun()


# ====================================================
# ALTERNATIVE: If you want fancy logout
# ====================================================

def logout():
    """
    Logout function with confirmation
    """
    import streamlit as st
    
    # Clear all session state
    for key in list(st.session_state.keys()):
        try:
            del st.session_state[key]
        except:
            pass
    
    # Show message and rerun
    st.success("✅ Logged out successfully!")
    import time
    time.sleep(1)
    st.rerun()





def render_sidebar():
    """Render modern sidebar with clean design"""
    apply_custom_sidebar_css()

    # Logo
    if st.session_state.get('logo_url', ''):
        st.sidebar.image(st.session_state.logo_url, width=120)

    # Pending Contact Requests - subtle badge (no blinking)
    try:
        client = get_google_sheets_client()
        SHEET_ID = get_agency_sheet_id()
        if client and SHEET_ID:
            cr_sheet = client.open_by_key(SHEET_ID).worksheet("Contact_Requests")
            cr_data = cr_sheet.get_all_records()
            pending_count = sum(
                1 for row in cr_data
                if str(row.get('Status', '')).strip().lower() == 'pending'
            )
            if pending_count > 0:
                st.sidebar.markdown(f"""
                <div style="
                    background: #FEF2F2;
                    color: #EF4444;
                    padding: 8px 14px;
                    border-radius: 8px;
                    font-weight: 600;
                    font-size: 13px;
                    border: 1px solid #FECACA;
                    margin: 0 8px 12px;
                    display: flex;
                    align-items: center;
                    gap: 8px;
                    font-family: 'Inter', sans-serif;
                ">
                    <span style="width:8px; height:8px; background:#EF4444; border-radius:50%; display:inline-block;"></span>
                    {pending_count} Pending Request{'s' if pending_count > 1 else ''}
                </div>
                """, unsafe_allow_html=True)
    except:
        pass

    agency_name = st.session_state.get('agency_name', 'Placify')

    # Agency name header
    st.sidebar.markdown(f"""
    <div style="padding: 8px 16px 16px; margin-bottom: 4px;">
        <h1 style="margin: 0; font-size: 18px; color: #0F172A !important; font-weight: 800; font-family: 'Inter', sans-serif; letter-spacing: -0.02em;">
            {agency_name}
        </h1>
    </div>
    """, unsafe_allow_html=True)

    # User info card
    if st.session_state.get("full_name"):
        full_name = st.session_state.get("full_name", "")
        email = st.session_state.get("email", "")
        role = (st.session_state.get("role") or "User").upper()

        # Get initials for avatar
        initials = "".join([n[0] for n in full_name.split()[:2]]).upper() if full_name else "U"

        st.sidebar.markdown(f"""
        <div style="margin: 0 8px 12px; padding: 12px; background: #F8FAFC; border-radius: 10px; border: 1px solid #E2E8F0;">
            <div style="display: flex; align-items: center; gap: 10px;">
                <div style="
                    width: 36px; height: 36px; background: linear-gradient(135deg, #6366F1, #8B5CF6);
                    border-radius: 10px; display: flex; align-items: center; justify-content: center;
                    color: white !important; font-size: 13px; font-weight: 700; flex-shrink: 0;
                    font-family: 'Inter', sans-serif;
                ">{initials}</div>
                <div style="min-width: 0;">
                    <div style="font-weight: 600; font-size: 13px; color: #0F172A !important; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; font-family: 'Inter', sans-serif;">{full_name}</div>
                    <div style="font-size: 11px; color: #94A3B8 !important; font-family: 'Inter', sans-serif;">{role}</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    st.sidebar.markdown("---")

    role = (st.session_state.get("role") or "").lower()

    if role == "admin":
        menu_items = ["Dashboard", "Company Management", "Vacancy Management", "Candidate Management",
                     "Advanced Filtering", "Job Matching", "Interview Management", "Reports & Analytics","Settings"]
    else:
        menu_items = ["Dashboard", "My Profile", "Applications", "Settings"]

    selected_menu = st.sidebar.radio(label="Navigation", options=menu_items, key="main_menu", label_visibility="collapsed")
    st.sidebar.markdown("---")
    logout_clicked = st.sidebar.button("Logout", use_container_width=True, key="logout_btn")

    return selected_menu, logout_clicked, role

# ==================================================== 
# GOOGLE SHEETS
# ==================================================== 
SHEET_ID = "1rpuXdpfwjy0BQcaZcn0Acbh-Se6L3PvyNGiNu4NLcPA"

if os.path.exists("credentials.json"):
    CRED_FILE = "credentials.json"
else:
    CRED_FILE = None

@st.cache_resource
def get_google_sheets_client():
    try:
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        if CRED_FILE:
            credentials = ServiceAccountCredentials.from_json_keyfile_name(CRED_FILE, scope)
        else:
            credentials = ServiceAccountCredentials.from_json_keyfile_dict(st.secrets["gcp_service_account"], scope)
        return gspread.authorize(credentials)
    except Exception as e:
        st.error(f"❌ Google Sheets error: {e}")
        return None

def get_agency_sheet_id():
    sheet_url = st.session_state.get("agency_sheet_url", "")
    if sheet_url:
        try:
            return sheet_url.split('/d/')[1].split('/')[0]
        except:
            return None
    return None

def _to_str_df(data):
    df = pd.DataFrame(data)
    if not df.empty:
        for col in df.columns:
            df[col] = df[col].astype(str)
    return df

def get_column_letter(n):              # 🆕 ADD THIS
    """Convert column number to Excel letter (1=A, 27=AA)"""
    result = ""
    while n > 0:
        n -= 1
        result = chr(65 + n % 26) + result
        n //= 26
    return result                       # 🆕 END
   

@st.cache_data(ttl=300)
def get_companies():
    try:
        client = get_google_sheets_client()
        if client:
            SHEET_ID = get_agency_sheet_id()
            if not SHEET_ID:
                return pd.DataFrame()
            sheet = client.open_by_key(SHEET_ID).worksheet("CID")
            return _to_str_df(sheet.get_all_records())
        return pd.DataFrame()
    except:
        return pd.DataFrame()

@st.cache_data(ttl=300)
def get_vacancies():
    try:
        client = get_google_sheets_client()
        if client:
            SHEET_ID = get_agency_sheet_id()
            if not SHEET_ID:
                return pd.DataFrame()
            sheet = client.open_by_key(SHEET_ID).worksheet("Sheet4")
            return _to_str_df(sheet.get_all_records())
        return pd.DataFrame()
    except:
        return pd.DataFrame()

@st.cache_data(ttl=300)
def get_candidates():
    try:
        client = get_google_sheets_client()
        if client:
            SHEET_ID = get_agency_sheet_id()
            if not SHEET_ID:
                return pd.DataFrame()
            sheet = client.open_by_key(SHEET_ID).worksheet("Candidates")
            return _to_str_df(sheet.get_all_records())
        return pd.DataFrame()
    except:
        return pd.DataFrame()

@st.cache_data(ttl=300)
def get_interviews():
    try:
        client = get_google_sheets_client()
        if client:
            SHEET_ID = get_agency_sheet_id()
            if not SHEET_ID:
                return pd.DataFrame()
            sheet = client.open_by_key(SHEET_ID).worksheet("Interview_Records")
            return _to_str_df(sheet.get_all_records())
        return pd.DataFrame()
    except:
        return pd.DataFrame()

def add_to_sheet(sheet_name, data_dict):
    try:
        client = get_google_sheets_client()
        if client:
            SHEET_ID = get_agency_sheet_id()
            if not SHEET_ID:
                return False
            sheet = client.open_by_key(SHEET_ID).worksheet(sheet_name)
            headers = sheet.row_values(1)
            row = [data_dict.get(h.strip(), "") for h in headers]
            sheet.append_row(row)
            st.success("✅ Data added!")
            st.cache_data.clear()
            return True
        return False
    except Exception as e:
        st.error(f"❌ Error: {e}")
        return False

def generate_next_cid():
    try:
        agency_code = st.session_state.get("agency_code", "AG000")
        companies_df = get_companies()
        if len(companies_df) == 0 or "CID" not in companies_df.columns:
            return f"{agency_code}CID0001"
        existing_cids = companies_df["CID"].tolist()
        numbers = []
        for cid in existing_cids:
            if isinstance(cid, str) and "CID" in cid:
                try:
                    numbers.append(int(cid.split("CID")[-1]))
                except:
                    pass
        next_num = max(numbers) + 1 if numbers else 1
        return f"{agency_code}CID{next_num:04d}"
    except:
        return f"{st.session_state.get('agency_code', 'AG000')}CID{pd.Timestamp.now().strftime('%m%d%H%M')}"

# SESSION STATE
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "username" not in st.session_state:
    st.session_state.username = None
if "role" not in st.session_state:
    st.session_state.role = None
if "full_name" not in st.session_state:
    st.session_state.full_name = ""
if "email" not in st.session_state:
    st.session_state.email = ""
if "active_interview_tab" not in st.session_state:
    st.session_state.active_interview_tab = 0  # Default tab index

# ==================================================== 
# DASHBOARD
# ==================================================== 
def admin_dashboard():
    render_breadcrumb(["Home", "Dashboard"])
    render_page_header("Dashboard Overview", "Track your agency's key metrics and recent activity")

    companies_df = get_companies()
    vacancies_df = get_vacancies()
    candidates_df = get_candidates()
    interviews_df = get_interviews()

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        render_kpi_card(title="Companies", value=len(companies_df), icon="🏢", color="primary")
    with col2:
        render_kpi_card(title="Vacancies", value=len(vacancies_df), icon="💼", color="info")
    with col3:
        render_kpi_card(title="Candidates", value=len(candidates_df), icon="👥", color="success")
    with col4:
        selected_count = len(interviews_df[interviews_df['Result Status'] == 'Selected']) if len(interviews_df) > 0 and 'Result Status' in interviews_df.columns else 0
        render_kpi_card(title="Selections", value=selected_count, icon="✓", color="warning")

    st.markdown("<div style='height:1.5rem;'></div>", unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown(f"""
        <div style="font-size:15px; font-weight:700; color:{COLORS['text_dark']}; margin-bottom:12px; font-family:'Inter',sans-serif;">
            Recent Interviews
        </div>
        """, unsafe_allow_html=True)
        if len(interviews_df) > 0:
            display_cols = ['Record ID', 'Full Name', 'Company Name', 'Job Title', 'Interview Status']
            available_cols = [c for c in display_cols if c in interviews_df.columns]
            render_enhanced_dataframe(interviews_df[available_cols].head(5))
        else:
            render_empty_state("No interviews yet")

    with col2:
        st.markdown(f"""
        <div style="font-size:15px; font-weight:700; color:{COLORS['text_dark']}; margin-bottom:12px; font-family:'Inter',sans-serif;">
            Interview Status
        </div>
        """, unsafe_allow_html=True)
        if len(interviews_df) > 0 and 'Interview Status' in interviews_df.columns:
            st.bar_chart(interviews_df['Interview Status'].value_counts())
        else:
            render_empty_state("No data available")

# ==================================================== 
# COMPANY MANAGEMENT
# ==================================================== 
def admin_company_mgmt():
    render_page_header("Company Management", "Add, view, and manage your registered companies")
    tab1, tab2, tab3 = st.tabs(["View All", "Add Company", "Edit/Delete"])
    
    with tab1:
        st.write("### All Companies")
        companies_df = get_companies()
        if len(companies_df) > 0:
            st.dataframe(companies_df, use_container_width=True)
        else:
            st.info("No companies found")
    
    with tab2:
        st.write("### Add New Company")
        with st.form("add_company_form"):
            col1, col2 = st.columns(2)
            with col1:
                name = st.text_input("Company Name *")
                industry = st.selectbox("Industry *", ["IT/Software", "Finance/Banking", "Healthcare/Medical", "Education/Training", "Manufacturing", "Other"])
                contact = st.text_input("Contact Email")
                phone = st.text_input("Contact Number")
            with col2:
                description = st.text_area("Description")
                address = st.text_area("Address")
                city = st.text_input("City")
                state = st.text_input("State")
            submitted = st.form_submit_button("➕ Add Company")
        
        if submitted and name:
            final_cid = generate_next_cid()
            data = {"Company Name": name, "CID": final_cid, "Industry": industry, "Contact Number": phone,
                   "Address of Company": address, "City": city, "State": state, "Email": contact,
                   "Company Description": description, "Date Added": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")}
            if add_to_sheet("CID", data):
                st.success(f"✅ Company '{name}' added with CID: {final_cid}")
                st.balloons()
    
    with tab3:
        st.info("⚠️ Edit directly in Google Sheets")
        if len(get_companies()) > 0:
            st.dataframe(get_companies(), use_container_width=True)

# ==================================================== 
# COMMON HELPERS FOR VACANCIES
# ==================================================== 
def _norm(s: str) -> str:
    return str(s).strip().lower().replace("_", " ").replace("-", " ")

def normalize_series(vals):
    return sorted({str(x).strip() for x in vals if str(x).strip()})

@st.cache_data(ttl=300)
def get_company_name_options():
    df = get_companies()
    pick = "Company Name" if "Company Name" in df.columns else ("Company_Name" if "Company_Name" in df.columns else None)
    return normalize_series(df[pick].dropna().tolist()) if pick else []

@st.cache_data(ttl=300)
def get_designation_options():
    client = get_google_sheets_client()
    if not client:
        return []
    
    SHEET_ID = get_agency_sheet_id()
    if not SHEET_ID:
        return []
    
    ws = client.open_by_key(SHEET_ID).worksheet("Sheet2")
    rows = ws.get_all_records()
    df = pd.DataFrame(rows)
    return normalize_series(df["Designation"].dropna().tolist()) if "Designation" in df.columns else []

@st.cache_data(ttl=300)
def get_sheet2_df():
    try:
        client = get_google_sheets_client()
        if not client:
            return pd.DataFrame()
        
        SHEET_ID = get_agency_sheet_id()
        if not SHEET_ID:
            return pd.DataFrame()
        
        ws = client.open_by_key(SHEET_ID).worksheet("Sheet2")
        rows = ws.get_all_records()
        return pd.DataFrame(rows)
    except Exception:
        return pd.DataFrame()

def lookup_cid(company_name: str) -> str:
    df = get_companies()
    if df.empty:
        return ""
    name_col = "Company Name" if "Company Name" in df.columns else ("Company_Name" if "Company_Name" in df.columns else None)
    if not name_col or "CID" not in df.columns:
        return ""
    key = str(company_name).strip().lower()
    hit = df[df[name_col].astype(str).str.strip().str.lower() == key]
    return str(hit.iloc[0]["CID"]) if not hit.empty and "CID" in hit.columns else ""

def lookup_dgn_id(job_title: str) -> str:
    df2 = get_sheet2_df()
    if df2.empty:
        return ""
    des_col = "Designation" if "Designation" in df2.columns else None
    dgn_col = "DGN ID" if "DGN ID" in df2.columns else ("DGN_ID" if "DGN_ID" in df2.columns else None)
    if not des_col or not dgn_col:
        return ""
    key = str(job_title).strip().lower()
    hit = df2[df2[des_col].astype(str).str.strip().str.lower() == key]
    return str(hit.iloc[0][dgn_col]) if not hit.empty else ""

@st.cache_data(ttl=300)
def get_education_options():
    """Priority: 1) Sheet 'education' → column 'Academic Education' 2) Fallback: 'Sheet4' → 'Education Required' 3) Defaults"""
    try:
        client = get_google_sheets_client()
        if not client:
            return ["12th", "Diploma", "B.Sc", "B.Tech", "M.Sc", "MBA"]

        SHEET_ID = get_agency_sheet_id()
        if not SHEET_ID:
            return ["12th", "Diploma", "B.Sc", "B.Tech", "M.Sc", "MBA"]

        ss = client.open_by_key(SHEET_ID)
        titles = [ws.title for ws in ss.worksheets()]
        edu_title = next((t for t in titles if t.strip().lower() == "education"), None)

        if edu_title:
            try:
                ws = ss.worksheet(edu_title)
                rows = ws.get_all_records()
                if rows:
                    df = pd.DataFrame(rows)
                    if "Academic Education" in df.columns:
                        return normalize_series(df["Academic Education"].dropna().tolist())
            except gspread.exceptions.WorksheetNotFound:
                pass

        if "Sheet4" in titles:
            try:
                ws4 = ss.worksheet("Sheet4")
                rows4 = ws4.get_all_records()
                if rows4:
                    df4 = pd.DataFrame(rows4)
                    if "Education Required" in df4.columns:
                        return normalize_series(df4["Education Required"].dropna().tolist())
            except gspread.exceptions.WorksheetNotFound:
                pass

        return ["12th", "Diploma", "B.Sc", "B.Tech", "M.Sc", "MBA"]
    except Exception:
        return ["12th", "Diploma", "B.Sc", "B.Tech", "M.Sc", "MBA"]

def add_to_sheet_safe(sheet_name, data_dict):
    """Header-insensitive append"""
    try:
        client = get_google_sheets_client()
        if not client:
            st.error("❌ Cannot connect to Google Sheets")
            return False
        
        SHEET_ID = get_agency_sheet_id()
        if not SHEET_ID:
            st.error("❌ Sheet not configured!")
            return False
        
        ws = client.open_by_key(SHEET_ID).worksheet(sheet_name)
        headers = ws.row_values(1)
        norm_map = {_norm(k): (v.strip() if isinstance(v, str) else v) for k, v in data_dict.items()}
        row = [norm_map.get(_norm(h), "") for h in headers]
        ws.append_row(row)
        st.cache_data.clear()
        return True
    except Exception as e:
        st.error(f"❌ Error adding data: {e}")
        return False

# ==================================================== 
# VACANCY MANAGEMENT
# ==================================================== 
def get_designations():
    """Fetch designations from Sheet2 (Designation sheet)"""
    try:
        client = get_google_sheets_client()
        if client:
            SHEET_ID = get_agency_sheet_id()
            if SHEET_ID:
                sheet = client.open_by_key(SHEET_ID).worksheet("Sheet2")
                all_data = sheet.get_all_values()
                
                if all_data and len(all_data) > 1:
                    headers = all_data[0]
                    rows = all_data[1:]
                    designations_df = pd.DataFrame(rows, columns=headers)
                    return designations_df
    except Exception as e:
        st.error(f"Error fetching designations: {str(e)}")
    
    return pd.DataFrame()


# ====================================================
# COMMON HELPERS FOR VACANCIES
# ====================================================
def _norm(s: str) -> str:
    return str(s).strip().lower().replace("_", " ").replace("-", " ")


def normalize_series(vals):
    return sorted({str(x).strip() for x in vals if str(x).strip()})


@st.cache_data(ttl=300)
def get_company_name_options():
    df = get_companies()
    pick = (
        "Company Name"
        if "Company Name" in df.columns
        else ("Company_Name" if "Company_Name" in df.columns else None)
    )
    return normalize_series(df[pick].dropna().tolist()) if pick else []


@st.cache_data(ttl=300)
def get_designation_options():
    """Fetch designations from Sheet2"""
    client = get_google_sheets_client()
    if not client:
        return []
    try:
        ws = client.open_by_key(SHEET_ID).worksheet("Sheet2")
        rows = ws.get_all_records()
        df = pd.DataFrame(rows)
        return (
            normalize_series(df["Designation"].dropna().tolist())
            if "Designation" in df.columns
            else []
        )
    except Exception:
        return []


@st.cache_data(ttl=300)
def get_sheet2_df():
    """Get full Sheet2 dataframe for DGN ID lookup"""
    try:
        client = get_google_sheets_client()
        if not client:
            return pd.DataFrame()
        ws = client.open_by_key(SHEET_ID).worksheet("Sheet2")
        rows = ws.get_all_records()
        return pd.DataFrame(rows)
    except Exception:
        return pd.DataFrame()


def lookup_cid(company_name: str) -> str:
    """Lookup CID from company name"""
    df = get_companies()
    if df.empty:
        return ""
    name_col = (
        "Company Name"
        if "Company Name" in df.columns
        else ("Company_Name" if "Company_Name" in df.columns else None)
    )
    if not name_col or "CID" not in df.columns:
        return ""
    key = str(company_name).strip().lower()
    hit = df[df[name_col].astype(str).str.strip().str.lower() == key]
    return str(hit.iloc[0]["CID"]) if not hit.empty and "CID" in hit.columns else ""


def lookup_dgn_id(job_title: str) -> str:
    """Lookup DGN ID from job title (designation) in Sheet2"""
    df2 = get_sheet2_df()
    if df2.empty:
        return ""
    des_col = "Designation" if "Designation" in df2.columns else None
    dgn_col = (
        "DGN ID"
        if "DGN ID" in df2.columns
        else ("DGN_ID" if "DGN_ID" in df2.columns else None)
    )
    if not des_col or not dgn_col:
        return ""
    key = str(job_title).strip().lower()
    hit = df2[df2[des_col].astype(str).str.strip().str.lower() == key]
    return str(hit.iloc[0][dgn_col]) if not hit.empty else ""


@st.cache_data(ttl=300)
def get_education_options():
    """
    Priority:
    1) Sheet title case-insensitive: 'education' → column 'Academic Education'
    2) Fallback: 'Sheet4' → column 'Education Required'
    3) Final defaults list
    """
    try:
        client = get_google_sheets_client()
        if not client:
            return ["12th", "Diploma", "B.Sc", "B.Tech", "M.Sc", "MBA"]

        ss = client.open_by_key(SHEET_ID)
        titles = [ws.title for ws in ss.worksheets()]
        edu_title = next(
            (t for t in titles if t.strip().lower() == "education"), None
        )

        if edu_title:
            try:
                ws = ss.worksheet(edu_title)
                rows = ws.get_all_records()
                if rows:
                    df = pd.DataFrame(rows)
                    if "Academic Education" in df.columns:
                        return normalize_series(
                            df["Academic Education"].dropna().tolist()
                        )
            except:
                pass

        if "Sheet4" in titles:
            try:
                ws4 = ss.worksheet("Sheet4")
                rows4 = ws4.get_all_records()
                if rows4:
                    df4 = pd.DataFrame(rows4)
                    if "Education Required" in df4.columns:
                        return normalize_series(
                            df4["Education Required"].dropna().tolist()
                        )
            except:
                pass

        return ["12th", "Diploma", "B.Sc", "B.Tech", "M.Sc", "MBA"]
    except Exception:
        return ["12th", "Diploma", "B.Sc", "B.Tech", "M.Sc", "MBA"]


# ====================================================
# ADMIN: VACANCY MANAGEMENT (COMPLETE VERSION)
# ====================================================
def admin_vacancy_mgmt():
    """
    COMPLETE Vacancy Management System:
    - Dashboard with metrics
    - View All Vacancies with filters
    - Add Vacancy (3-tab form with Sheet2 dropdown + Fixed row detection)
    - Edit Vacancy
    """
    render_page_header("Vacancy Management", "Create and manage job openings across your companies")
    
    # Fetch data once
    vacancies_df = get_vacancies()
    companies_df = get_companies()
    
    # Create 4 tabs
    tab1, tab2, tab3, tab4 = st.tabs([
        "📊 Dashboard",
        "📋 View All Vacancies", 
        "➕ Add Vacancy",
        "✏️ Edit Vacancy"
    ])

    # ========== TAB 1: DASHBOARD ==========
    with tab1:
        st.markdown("### 📊 Vacancy Dashboard")
        
        if len(vacancies_df) > 0:
            # Find status column (case-insensitive)
            status_col = None
            for col in vacancies_df.columns:
                if col.lower() == 'status':
                    status_col = col
                    break
            
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                if status_col:
                    open_count = len(vacancies_df[vacancies_df[status_col].str.strip().str.upper() == 'OPEN'])
                else:
                    open_count = 0
                st.metric("🟢 Open", open_count)
            
            with col2:
                if status_col:
                    closed_count = len(vacancies_df[vacancies_df[status_col].str.strip().str.upper() == 'CLOSED'])
                else:
                    closed_count = 0
                st.metric("🔴 Closed", closed_count)
            
            with col3:
                total_count = len(vacancies_df)
                st.metric("📊 Total", total_count)
            
            with col4:
                if 'Vacancy Filled' in vacancies_df.columns and 'Vacancy Count' in vacancies_df.columns:
                    try:
                        vacancies_df['Vacancy Filled'] = pd.to_numeric(vacancies_df['Vacancy Filled'], errors='coerce').fillna(0)
                        vacancies_df['Vacancy Count'] = pd.to_numeric(vacancies_df['Vacancy Count'], errors='coerce').fillna(0)
                        filled_count = int(vacancies_df['Vacancy Filled'].sum())
                    except:
                        filled_count = 0
                else:
                    filled_count = 0
                st.metric("✅ Filled", filled_count)
            
            st.markdown("---")
            
            # Recent Vacancies
            if len(vacancies_df) > 0:
                st.markdown("### 🆕 Recent Vacancies")
                recent = vacancies_df.head(10)
                display_cols = ['Company Name', 'Job Title', 'Vacancy Count', 'Salary', 'Status']
                available_cols = [col for col in display_cols if col in recent.columns]
                st.dataframe(recent[available_cols], use_container_width=True, hide_index=True)
            
            st.markdown("---")
            
            # Urgent Vacancies
            if 'Urgency Level' in vacancies_df.columns:
                urgent = vacancies_df[
                    vacancies_df['Urgency Level'].str.strip().str.upper().isin(['HIGH', 'CRITICAL'])
                ]
                if len(urgent) > 0:
                    st.markdown("### ⚠️ Urgent Vacancies")
                    st.warning(f"{len(urgent)} urgent vacancies need attention!")
                    display_cols = ['Company Name', 'Job Title', 'Urgency Level', 'Date Added']
                    available_cols = [col for col in display_cols if col in urgent.columns]
                    st.dataframe(urgent[available_cols], use_container_width=True, hide_index=True)
        else:
            st.info("No vacancies found. Add your first vacancy!")

    # ========== TAB 2: VIEW ALL VACANCIES ==========
    with tab2:
        st.markdown("### 📋 All Vacancies")
        
        if len(vacancies_df) > 0:
            # Filters
            col1, col2, col3 = st.columns(3)
            
            # Find status column
            status_col = None
            for col in vacancies_df.columns:
                if col.lower() == 'status':
                    status_col = col
                    break
            
            with col1:
                if status_col:
                    status_filter = st.multiselect(
                        "Filter by Status",
                        options=vacancies_df[status_col].unique().tolist(),
                        default=vacancies_df[status_col].unique().tolist()
                    )
                else:
                    status_filter = []
            
            with col2:
                if 'Company Name' in vacancies_df.columns:
                    company_filter = st.multiselect(
                        "Filter by Company",
                        options=vacancies_df['Company Name'].unique().tolist(),
                        default=[]
                    )
                else:
                    company_filter = []
            
            with col3:
                search_text = st.text_input("Search (Job Title)")
            
            # Apply filters
            filtered_df = vacancies_df.copy()
            
            if status_filter and status_col:
                filtered_df = filtered_df[filtered_df[status_col].isin(status_filter)]
            
            if company_filter and 'Company Name' in vacancies_df.columns:
                filtered_df = filtered_df[filtered_df['Company Name'].isin(company_filter)]
            
            if search_text:
                if 'Job Title' in filtered_df.columns:
                    filtered_df = filtered_df[
                        filtered_df['Job Title'].str.contains(search_text, case=False, na=False)
                    ]
            
            st.write(f"**Showing {len(filtered_df)} / {len(vacancies_df)} vacancies**")
            st.dataframe(filtered_df, use_container_width=True, height=400)
            
            # Download button
            csv = filtered_df.to_csv(index=False)
            st.download_button(
                label="📥 Download Filtered Data (CSV)",
                data=csv,
                file_name=f"vacancies_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv"
            )
        else:
            st.info("No vacancies found")

    # ========== TAB 3: ADD VACANCY (3-TAB FORM) ==========
    with tab3:
        st.markdown("### ➕ Add New Vacancy")

        # Fetch dropdown options
        company_opts = get_company_name_options()
        dgn_opts = get_designation_options()
        edu_opts = get_education_options()

        if not company_opts:
            st.warning("⚠️ No companies found. Please add companies first!")
            return

        # 3-TAB FORM
        with st.form("add_vacancy_form", clear_on_submit=True):
            # Create 3 inner tabs for form sections
            t_basic, t_req, t_log = st.tabs(["Basic", "Requirements", "Logistics"])

            # ========== BASIC TAB ==========
            with t_basic:
                col1, col2 = st.columns(2)
                with col1:
                    company_name = st.selectbox(
                        "Company Name *",
                        company_opts,
                        index=0 if company_opts else None,
                        placeholder="Select company",
                    )
                    job_title = st.selectbox(
                        "Job Title (Designation) *",
                        dgn_opts,
                        index=0 if dgn_opts else None,
                        placeholder="Select designation",
                        help="Designation from Sheet2"
                    )
                    vacancy_count = st.number_input(
                        "Vacancy Count *", min_value=1, step=1, value=1
                    )
                with col2:
                    salary = st.text_input("Salary *", placeholder="e.g., 3-5 LPA")
                    job_desc = st.text_area("Job Description", height=120)

            # ========== REQUIREMENTS TAB ==========
            with t_req:
                col1, col2 = st.columns(2)
                with col1:
                    edu_req = st.selectbox(
                        "Education Required *",
                        edu_opts if edu_opts else ["12th", "Diploma", "B.Sc", "B.Tech", "M.Sc", "MBA"],
                        index=0,
                        placeholder="Select education",
                    )
                    skills_req = st.text_input("Skills Required", placeholder="e.g., Python, SQL")
                    exp_req = st.text_input("Experience Required", placeholder="e.g., 2-5 years")
                    gender_pref = st.selectbox(
                        "Gender Preference", ["Any", "Male", "Female", "Other"]
                    )
                with col2:
                    urgency = st.selectbox(
                        "Urgency Level", ["Low", "Medium", "High", "Critical"], index=0
                    )
                    age_min = st.number_input(
                        "Age Range Min", min_value=18, max_value=100, value=21, step=1
                    )
                    age_max = st.number_input(
                        "Age Range Max", min_value=18, max_value=100, value=60, step=1
                    )
                    pref_loc = st.text_input("Preferred Candidate Location", placeholder="e.g., Mumbai, Pune")

            # ========== LOGISTICS TAB ==========
            with t_log:
                col1, col2 = st.columns(2)
                with col1:
                    job_city = st.text_input("Job Location/City *", placeholder="e.g., Mumbai")
                    job_type = st.selectbox(
                        "Job Type *",
                        ["Full-time", "Part-time", "Contract", "Internship"],
                    )
                    work_mode = st.selectbox(
                        "Work Mode *", ["On-site", "Remote", "Hybrid"]
                    )
                    job_timing = st.text_input("Job Timing", placeholder="e.g., 9 AM - 6 PM")
                with col2:
                    shift_timings = st.text_input("Shift Timings", placeholder="e.g., Day Shift")
                    notice_ok = st.selectbox(
                        "Notice Period Acceptable",
                        ["Any", "Immediate", "15 days", "30 days", "60 days"],
                    )
                    contact_person = st.text_input("Contact Person")
                    contact_number = st.text_input("Contact Number")
                
                notes = st.text_area("Additional Notes", height=100)

            submit_vac = st.form_submit_button("➕ Add Vacancy", type="primary")

        # ========== FORM SUBMISSION WITH FIXED ROW DETECTION ==========
        if submit_vac:
            # Validation
            if not str(company_name).strip() or not str(job_title).strip():
                st.error("⚠️ Company Name और Job Title आवश्यक हैं!")
            else:
                # Lookup CID and DGN ID
                cid_val = lookup_cid(company_name)
                dgn_id_val = lookup_dgn_id(job_title)

                try:
                    client = get_google_sheets_client()
                    if not client:
                        st.error("❌ Cannot connect to Google Sheets")
                        return
                    
                    # ========== FIXED ROW DETECTION ==========
                    sheet = client.open_by_key(SHEET_ID).worksheet("Sheet4")
                    all_data = sheet.get_all_values()
                    
                    if not all_data:
                        st.error("❌ Sheet4 is empty or not found")
                        return
                    
                    headers = all_data[0]
                    
                    # Find the ACTUAL last row with data
                    last_row_with_data = 1
                    for idx, row in enumerate(all_data[1:], start=2):
                        if any(cell.strip() for cell in row):
                            last_row_with_data = idx
                    
                    next_row = last_row_with_data + 1
                    
                    # Prepare data matching Sheet4 columns
                    new_vacancy_data = {
                        "Company Name": str(company_name).strip(),
                        "CID": cid_val,
                        "Job Title": str(job_title).strip(),
                        "DGN ID": dgn_id_val,
                        "Salary": str(salary).strip(),
                        "Job Description": str(job_desc).strip(),
                        "Education Required": str(edu_req).strip(),
                        "Skills Required": str(skills_req).strip(),
                        "Experience Required": str(exp_req).strip(),
                        "Vacancy Count": vacancy_count,
                        "Contact Person": str(contact_person).strip(),
                        "Contact Number": str(contact_number).strip(),
                        "Additional Notes": str(notes).strip(),
                        "Date Added": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "Job Location/City": str(job_city).strip(),
                        "Gender Preference": str(gender_pref).strip(),
                        "Job Type": str(job_type).strip(),
                        "Job Timing": str(job_timing).strip(),
                        "Shift Timings": str(shift_timings).strip(),
                        "Notice Period Acceptable": str(notice_ok).strip(),
                        "Work Mode": str(work_mode).strip(),
                        "Age Range Min": age_min,
                        "Age Range Max": age_max,
                        "Urgency Level": str(urgency).strip(),
                        "Preferred Candidate Location": str(pref_loc).strip(),
                        "Vacancy Filled": "0",
                        "Status": "Open",
                    }
                    
                    # Create row data in EXACT order of existing headers
                    row_data = []
                    for header in headers:
                        row_data.append(new_vacancy_data.get(header, ''))
                    
                    # ========== UPDATE TO EXACT ROW ==========
                    last_col = get_column_letter(len(headers))
                    range_to_update = f"A{next_row}:{last_col}{next_row}"
                    sheet.update(range_to_update, [row_data])
                    
                    st.success(f"✅ Vacancy for '{job_title}' added successfully at row {next_row}!")
                    if dgn_id_val:
                        st.info(f"🆔 DGN ID: {dgn_id_val}")
                    st.balloons()
                    st.cache_data.clear()
                    
                    import time
                    time.sleep(2)
                    st.rerun()
                    
                except Exception as e:
                    st.error(f"❌ Error adding vacancy: {str(e)}")
                    import traceback
                    st.code(traceback.format_exc())

    # ========== TAB 4: EDIT VACANCY ==========
    with tab4:
        st.markdown("### ✏️ Edit Vacancy")
        
        if len(vacancies_df) > 0:
            # Select vacancy to edit
            vacancy_options = vacancies_df.apply(
                lambda x: f"{x.get('Company Name', 'N/A')} - {x.get('Job Title', 'N/A')} ({x.get('Status', 'N/A')})", 
                axis=1
            ).tolist()
            
            selected_vacancy = st.selectbox(
                "Select Vacancy to Edit",
                vacancy_options,
                key="edit_select_vacancy"
            )
            
            if selected_vacancy:
                vacancy_idx = vacancy_options.index(selected_vacancy)
                vacancy_data = vacancies_df.iloc[vacancy_idx]
                
                st.markdown("---")
                
                # Fetch options
                company_opts = get_company_name_options()
                dgn_opts = get_designation_options()
                edu_opts = get_education_options()
                
                # Show current info
                st.info(f"Editing: {vacancy_data.get('Company Name', 'N/A')} - {vacancy_data.get('Job Title', 'N/A')}")
                
                with st.form("edit_vacancy_form"):
                    st.markdown("#### 📝 Edit Vacancy Details")
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        # Company (read-only for edit)
                        st.text_input("Company Name", value=vacancy_data.get('Company Name', ''), disabled=True)
                        
                        # Job Title dropdown
                        current_job = vacancy_data.get('Job Title', '')
                        job_idx = dgn_opts.index(current_job) if current_job in dgn_opts else 0
                        job_title = st.selectbox(
                            "Job Title (Designation) *",
                            dgn_opts,
                            index=job_idx,
                        )
                        
                        vacancy_count = st.number_input(
                            "Vacancy Count *", 
                            min_value=1, 
                            value=int(vacancy_data.get('Vacancy Count', 1)), 
                            step=1
                        )
                        
                        salary = st.text_input("Salary *", value=vacancy_data.get('Salary', ''))
                    
                    with col2:
                        # Education dropdown
                        current_edu = vacancy_data.get('Education Required', '')
                        edu_idx = edu_opts.index(current_edu) if current_edu in edu_opts else 0
                        edu_req = st.selectbox(
                            "Education Required *",
                            edu_opts,
                            index=edu_idx,
                        )
                        
                        exp_req = st.text_input("Experience Required", value=vacancy_data.get('Experience Required', ''))
                        
                        # Find status column
                        status_col = None
                        for col in vacancies_df.columns:
                            if col.lower() == 'status':
                                status_col = col
                                break
                        
                        current_status = vacancy_data.get(status_col, 'Open') if status_col else 'Open'
                        status = st.selectbox(
                            "Status *", 
                            ["Open", "Closed"],
                            index=["Open", "Closed"].index(current_status) if current_status in ["Open", "Closed"] else 0
                        )
                        
                        vacancy_filled = st.number_input(
                            "Vacancy Filled", 
                            min_value=0, 
                            value=int(vacancy_data.get('Vacancy Filled', 0)), 
                            step=1
                        )
                    
                    job_desc = st.text_area("Job Description", value=vacancy_data.get('Job Description', ''), height=100)
                    skills_req = st.text_input("Skills Required", value=vacancy_data.get('Skills Required', ''))
                    
                    update_button = st.form_submit_button("💾 Update Vacancy", type="primary")
                
                if update_button:
                    try:
                        client = get_google_sheets_client()
                        if not client:
                            st.error("❌ Cannot connect to Google Sheets")
                            return
                        
                        sheet = client.open_by_key(SHEET_ID).worksheet("Sheet4")
                        all_data = sheet.get_all_values()
                        
                        if not all_data:
                            st.error("Sheet4 is empty.")
                            return
                        
                        headers = all_data[0]
                        
                        # Row to update is vacancy_idx + 2 (1 for header, 1 for 0-indexing)
                        row_to_update = vacancy_idx + 2
                        
                        # Get new DGN ID
                        dgn_id_val = lookup_dgn_id(job_title)
                        
                        # Prepare update data
                        update_data = {
                            'Job Title': job_title,
                            'DGN ID': dgn_id_val,
                            'Salary': salary,
                            'Vacancy Count': str(vacancy_count),
                            'Experience Required': exp_req,
                            'Education Required': edu_req,
                            'Job Description': job_desc,
                            'Skills Required': skills_req,
                            'Vacancy Filled': str(vacancy_filled),
                            'Status': status
                        }
                        
                        updates = []
                        for field, value in update_data.items():
                            if field in headers:
                                col_idx = headers.index(field) + 1
                                col_letter = chr(64 + col_idx)
                                updates.append({
                                    'range': f"{col_letter}{row_to_update}",
                                    'values': [[value]]
                                })
                        
                        sheet.batch_update(updates)
                        
                        st.success(f"✅ Vacancy updated successfully!")
                        st.cache_data.clear()
                        
                        import time
                        time.sleep(2)
                        st.rerun()
                        
                    except Exception as e:
                        st.error(f"❌ Error updating vacancy: {str(e)}")
        else:
            st.info("No vacancies found to edit")
# ==================================================== 
# CANDIDATE MANAGEMENT
# ==================================================== 
def admin_candidate_mgmt():
    render_page_header("Candidate Management", "Register and manage candidate profiles")
    tab1, tab2, tab3 = st.tabs(["View All", "Quick Add", "Full Form"])
    
    with tab1:
        st.write("### All Candidates")
        st.info("💡 Use 'Advanced Filtering' for detailed filters")
        candidates_df = get_candidates()
        if len(candidates_df) > 0:
            st.dataframe(candidates_df, use_container_width=True, height=400)
            csv = candidates_df.to_csv(index=False)
            st.download_button("📤 Download CSV", data=csv, file_name="candidates.csv", mime="text/csv")
        else:
            st.info("No candidates found")
    
    with tab2:
        st.write("### Quick Add")
        st.info("Quick form - Full implementation available")
    
    with tab3:
        st.write("### Full Wizard Form")
        render_wizard()

# ==================================================== 
# ADVANCED FILTERING
# ==================================================== 
def admin_advanced_filtering():
    render_page_header("Advanced Filtering", "Filter candidates and companies with 70+ columns available")
    
    tab1, tab2 = st.tabs(["Filter Candidates", "Filter Companies"])
    with tab1:
        render_candidate_filter()
    with tab2:
        render_company_filter()

# ==================================================== 
# JOB MATCHING
# ==================================================== 
def admin_job_matching():
    render_page_header("Job Matching Engine", "AI-powered candidate-vacancy matching with weighted scoring")
    
    # 1) Data load – Advanced Filtering ka respect
    col1, col2 = st.columns(2)
    
    # Candidates
    with col1:
        st.markdown("### 👥 Candidates Data")
        if "filtered_df" in st.session_state and st.session_state.get("filtered_df") is not None:
            candidates_df = st.session_state["filtered_df"]
            st.success(f"Using {len(candidates_df)} filtered candidates from Advanced Filtering.")
        else:
            candidates_df = get_candidates()
            st.warning(f"No candidate filters applied. Using all {len(candidates_df)} candidates.")
    
    # Companies / Vacancies
    with col2:
        st.markdown("### 🏢 Companies Data")
        if "companies_filtered_df" in st.session_state and st.session_state.get("companies_filtered_df") is not None:
            vacancies_df = st.session_state["companies_filtered_df"]
            st.success(f"Using {len(vacancies_df)} filtered companies from Advanced Filtering.")
        else:
            vacancies_df = get_vacancies()
            st.warning(f"No company filters applied. Using all {len(vacancies_df)} vacancies.")
    
    # Safety checks
    if len(candidates_df) == 0:
        st.error("No candidates available for matching!")
        return
    if len(vacancies_df) == 0:
        st.error("No companies/vacancies available for matching!")
        return
    
    st.markdown(f"**Using:** {len(candidates_df)} candidates × {len(vacancies_df)} vacancies")
    st.markdown("---")
    
    with st.expander("How Smart Matching Works", expanded=False):
        st.markdown("""
- **Critical fields (100%)**
  - Job Title – 40% weight (checks all 3 job preferences)
  - Location – 30% weight (preferred + current city)
  - Salary – 30% weight (numeric with 30% tolerance)

- **Optional bonus (20%)**
  - Skills
  - Education
  - Experience

- **Thresholds**
  - Minimum field match: 50%
  - Minimum total score: 40%
  - Uses fuzzy matching for text fields
        """)
    
    st.markdown("---")
    
    # 2) Controls row
    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        run_matching_btn = st.button("Run Smart Matching", type="primary", use_container_width=True)
    with col2:
        refresh_btn = st.button("Refresh Data", use_container_width=True)
    with col3:
        clear_btn = st.button("Clear Matches", use_container_width=True)
    
    if refresh_btn:
        st.cache_data.clear()
        st.success("Data refreshed from Google Sheets.")
        st.rerun()
    
    if clear_btn:
        if "matches_admin" in st.session_state:
            del st.session_state["matches_admin"]
        st.success("Cleared in-memory matches.")
        return
    
    # 3) Run matching
    if run_matching_btn:
        progress_placeholder = st.empty()
        status_placeholder = st.empty()
        
        def _progress(p):
            progress_placeholder.progress(p)
        
        def _status(txt):
            status_placeholder.text(txt)
        
        with st.spinner("Running matching algorithm..."):
            matches_df = run_matching(
                candidates_df,
                vacancies_df,
                progress_callback=_progress,
                status_callback=_status,
            )
            st.session_state["matches_admin"] = matches_df
        
        progress_placeholder.empty()
        status_placeholder.empty()
    
    # 4) Show results
    if "matches_admin" in st.session_state and len(st.session_state["matches_admin"]) > 0:
        matches_df = st.session_state["matches_admin"]
        st.success(f"✅ Found {len(matches_df)} matches.")
        st.markdown("---")
        st.subheader("Match Results")
        
        # Reset index and prepare selection list
        matches_df = matches_df.reset_index(drop=True)
        selected_rows = []
        
        # -------- Export controls (top) --------
        st.markdown("### 📤 Export Controls")
        col_e1, col_e2, col_e3 = st.columns([2, 1, 1])
        
        with col_e1:
            st.info(f"Selected: {len(selected_rows)} / {len(matches_df)} matches")
        
        with col_e2:
            if st.button("Export ALL Matches to Interview Records", key="adm_export_all_top"):
                gc = get_google_sheets_client()
                if gc:
                    SHEET_ID = get_agency_sheet_id()
                    if not SHEET_ID:
                        st.error("❌ Sheet not configured! Please login again.")
                        return
                    
                    all_matches = [row.to_dict() for _, row in matches_df.iterrows()]
                    success, msg = export_to_interview_sheet(gc, SHEET_ID, all_matches)
                    if success:
                        st.success(msg)
                        st.balloons()
                    else:
                        st.error(msg)
                else:
                    st.error("Google Sheets connection failed.")
        
        with col_e3:
            pass
        
        st.markdown("---")
        
        # -------- Row-wise results with selection + Quick Add --------
        for idx, row in matches_df.iterrows():
            c1, c2, c3 = st.columns([0.08, 0.72, 0.20])
            
            with c1:
                selected = st.checkbox("", key=f"adm_match_sel_{idx}")
                if selected:
                    selected_rows.append(idx)
            
            with c2:
                st.markdown(f"**{row['Full Name']}** → **{row['Company Name']}** (CID: {row['CID']})")
                st.caption(
                    f"Job: {row['Job Title']} | "
                    f"Match Score: **{row['Match Score']}%** | "
                    f"Salary: {row.get('Salary', 'N/A')} | "
                    f"Industry: {row.get('Industry', 'N/A')}"
                )
                st.caption(f"Contact: {row.get('Contact', 'N/A')} ({row.get('Phone', 'N/A')})")
            
            with c3:
                if st.button("Quick Add", key=f"adm_quick_add_{idx}"):
                    gc = get_google_sheets_client()
                    if gc:
                        SHEET_ID = get_agency_sheet_id()
                        if not SHEET_ID:
                            st.error("❌ Sheet not configured! Please login again.")
                            return
                        
                        success, msg = export_to_interview_sheet(gc, SHEET_ID, [row.to_dict()])
                        if success:
                            st.success(msg)
                        else:
                            st.error(msg)
                    else:
                        st.error("Google Sheets connection failed.")
            
            st.markdown("---")
        
        # -------- Batch export (selected) --------
        st.markdown(f"**Selected: {len(selected_rows)} matches**")
        if st.button(
            f"Export Selected ({len(selected_rows)})",
            type="primary",
            key="adm_export_selected",
            disabled=len(selected_rows) == 0,
        ):
            if not selected_rows:
                st.warning("Please select at least one match.")
            else:
                gc = get_google_sheets_client()
                if gc:
                    SHEET_ID = get_agency_sheet_id()
                    if not SHEET_ID:
                        st.error("❌ Sheet not configured! Please login again.")
                        return
                    
                    selected_matches = [matches_df.iloc[i].to_dict() for i in selected_rows]
                    success, msg = export_to_interview_sheet(gc, SHEET_ID, selected_matches)
                    if success:
                        st.success(msg)
                        st.balloons()
                    else:
                        st.error(msg)
                else:
                    st.error("Google Sheets connection failed.")
    
    elif "matches_admin" in st.session_state and len(st.session_state["matches_admin"]) == 0:
        st.warning("⚠️ No Matches Found!")
        st.info("""
        💡 **Why no matches?**
        - Job preferences don't match available positions
        - Salary expectations are too high or too low
        - Location preferences don't match company locations
        - Experience or skills don't align with requirements
        
        **💡 What to try:**
        1. Adjust candidate filters (lower salary expectations, expand locations)
        2. Adjust company/vacancy filters (different industries, job types)
        3. Check candidate Job Preferences 1, 2, 3 in the sheet
        4. Run matching again with different filters
        """)
    
    else:
        st.info("Run Smart Matching to see results.")

# ==================================================== 
# INTERVIEW MANAGEMENT HELPER FUNCTIONS
# ==================================================== 
def check_existing_selections(candidate_id):
    """Check if candidate already has any 'Selected' result status"""
    try:
        client = get_google_sheets_client()
        if not client:
            return []
        
        SHEET_ID = get_agency_sheet_id()
        if not SHEET_ID:
            return []
        
        sheet = client.open_by_key(SHEET_ID).worksheet("Interview_Records")
        all_data = sheet.get_all_values()
        
        if len(all_data) <= 1:
            return []
        
        headers = all_data[0]
        candidate_id_col = headers.index('Candidate ID') if 'Candidate ID' in headers else -1
        result_status_col = headers.index('Result Status') if 'Result Status' in headers else -1
        record_id_col = headers.index('Record ID') if 'Record ID' in headers else -1
        company_col = headers.index('Company Name') if 'Company Name' in headers else -1
        job_title_col = headers.index('Job Title') if 'Job Title' in headers else -1
        
        if candidate_id_col == -1 or result_status_col == -1:
            return []
        
        existing_selections = []
        for row_idx, row in enumerate(all_data[1:], start=2):
            if (candidate_id_col < len(row) and 
                str(row[candidate_id_col]).strip() == str(candidate_id).strip() and
                result_status_col < len(row) and
                str(row[result_status_col]).strip() == "Selected"):
                
                record_id = row[record_id_col] if record_id_col < len(row) else "Unknown"
                company = row[company_col] if company_col < len(row) else "Unknown"
                job_title = row[job_title_col] if job_title_col < len(row) else "Unknown"
                
                existing_selections.append({
                    'row_num': row_idx,
                    'record_id': record_id,
                    'company': company,
                    'job_title': job_title
                })
        
        return existing_selections
    except Exception as e:
        return []

def update_selection_status(current_record_id, keep_selection, existing_selections):
    """Update selection status based on user choice"""
    try:
        client = get_google_sheets_client()
        if not client:
            return False
        
        SHEET_ID = get_agency_sheet_id()
        if not SHEET_ID:
            return False
        
        sheet = client.open_by_key(SHEET_ID).worksheet("Interview_Records")
        all_data = sheet.get_all_values()
        headers = all_data[0]
        
        result_status_col = headers.index('Result Status') + 1 if 'Result Status' in headers else -1
        record_id_col = headers.index('Record ID') if 'Record ID' in headers else -1
        
        if result_status_col == -1 or record_id_col == -1:
            return False
        
        if keep_selection == 'current':
            reject_rows = [sel['row_num'] for sel in existing_selections]
        else:
            reject_rows = []
            for row_idx, row in enumerate(all_data[1:], start=2):
                if (record_id_col < len(row) and 
                    str(row[record_id_col]).strip() == str(current_record_id).strip()):
                    reject_rows.append(row_idx)
        
        updates = []
        for row_num in reject_rows:
            updates.append({
                'range': f"{chr(64 + result_status_col)}{row_num}",
                'values': [['Rejected']]
            })
        
        if updates:
            sheet.batch_update(updates)
        
        return True
    except Exception as e:
        return False

def cancel_pending_entries(candidate_id, current_record_id):
    """Cancel all PENDING entries for a candidate when one is SELECTED"""
    try:
        client = get_google_sheets_client()
        if not client:
            return False
        
        SHEET_ID = get_agency_sheet_id()
        if not SHEET_ID:
            return False
        
        sheet = client.open_by_key(SHEET_ID).worksheet("Interview_Records")
        all_data = sheet.get_all_values()
        
        if len(all_data) <= 1:
            return False
        
        headers = all_data[0]
        candidate_id_col = headers.index('Candidate ID') if 'Candidate ID' in headers else -1
        result_status_col = headers.index('Result Status') if 'Result Status' in headers else -1
        record_id_col = headers.index('Record ID') if 'Record ID' in headers else -1
        
        if candidate_id_col == -1 or result_status_col == -1:
            return False
        
        pending_rows = []
        for row_idx, row in enumerate(all_data[1:], start=2):
            record_id = row[record_id_col] if record_id_col < len(row) else ""
            result_status = row[result_status_col] if result_status_col < len(row) else ""
            cand_id = row[candidate_id_col] if candidate_id_col < len(row) else ""
            
            if (str(cand_id).strip() == str(candidate_id).strip() and
                str(result_status).strip() == "Pending" and
                str(record_id).strip() != str(current_record_id).strip()):
                pending_rows.append(row_idx)
        
        if pending_rows:
            updates = []
            result_col = headers.index('Result Status') + 1 if 'Result Status' in headers else -1
            interview_status_col = headers.index('Interview Status') + 1 if 'Interview Status' in headers else -1
            
            for row_num in pending_rows:
                updates.append({
                    'range': f"{chr(64 + result_col)}{row_num}",
                    'values': [['Cancelled due to Selection']]
                })
                updates.append({
                    'range': f"{chr(64 + interview_status_col)}{row_num}",
                    'values': [['Cancelled due to Selection']]
                })
            
            if updates:
                sheet.batch_update(updates)
        
        return True
    except Exception as e:
        return False

def get_closed_vacancy_keys(vacancies_df):
    """Extract closed vacancies as (CID, Job Title) tuples"""
    if len(vacancies_df) == 0:
        return set()
    
    closed = vacancies_df[vacancies_df['status'].str.strip().str.upper() == 'CLOSED'].copy()
    if len(closed) == 0:
        return set()
    
    closed_keys = set()
    for _, row in closed.iterrows():
        cid = str(row.get('CID', '')).strip()
        job_title = str(row.get('Job Title', '')).strip()
        if cid and job_title:
            closed_keys.add((cid, job_title))
    
    return closed_keys

def is_vacancy_closed(interview_row, vacancies_df):
    """Check if this interview's vacancy is closed"""
    closed_keys = get_closed_vacancy_keys(vacancies_df)
    interview_cid = str(interview_row.get('CID', '')).strip()
    interview_job = str(interview_row.get('Job Title', '')).strip()
    return (interview_cid, interview_job) in closed_keys

def get_schedulable_interviews(interviews_df, vacancies_df):
    """Filter interviews that can be scheduled"""
    if len(interviews_df) == 0:
        return pd.DataFrame()
    
    matched = interviews_df[interviews_df['Interview Status'].str.strip() == 'Matched'].copy()
    if len(matched) == 0:
        return pd.DataFrame()
    
    matched = matched[matched['Result Status'] != 'Rejected']
    matched = matched[~matched.apply(lambda row: is_vacancy_closed(row, vacancies_df), axis=1)]
    
    grouped = interviews_df.groupby(['Candidate ID', 'Company Name', 'Job Title'])
    duplicates_to_hide = set()
    
    for (cand_id, company, job_title), group in grouped:
        if len(group) > 1:
            statuses = group['Interview Status'].unique()
            if 'Interview Scheduled' in statuses or 'Interview Completed' in statuses:
                matched_in_group = group[group['Interview Status'] == 'Matched']
                duplicates_to_hide.update(matched_in_group['Record ID'].tolist())
    
    schedulable = matched[~matched['Record ID'].isin(duplicates_to_hide)]
    return schedulable.reset_index(drop=True)

def get_updatable_interviews(interviews_df, vacancies_df):
    """Filter interviews that can have results updated"""
    if len(interviews_df) == 0:
        return pd.DataFrame()
    
    active = interviews_df[interviews_df['Interview Status'].isin(['Interview Scheduled', 'Interview Completed'])].copy()
    active = active[~active['Result Status'].isin(['Selected', 'Cancelled due to Selection'])]
    
    selected_candidates = interviews_df[interviews_df['Result Status'] == 'Selected']['Candidate ID'].unique()
    active = active[~active['Candidate ID'].isin(selected_candidates)]
    active = active[~active.apply(lambda row: is_vacancy_closed(row, vacancies_df), axis=1)]
    
    return active.reset_index(drop=True)

# ==================================================== 
# INTERVIEW MANAGEMENT - REFACTORED WITH RADIO + IF/ELSE
# ==================================================== 
def admin_interview_mgmt():
    """Interview Management with Radio Button Tab Control - NO RERUN ISSUES"""
    render_page_header("Interview Management", "Schedule, track, and manage interviews across all positions")
    
    # Fetch data once
    interviews_df = get_interviews()
    vacancies_df = get_vacancies()
    candidates_df = get_candidates()
    companies_df = get_companies()
    
    # ========== RADIO BUTTON TAB SELECTOR ==========
    selected_tab = st.radio(
        "Select Section:",
        ["📊 Dashboard", "🗓️ Schedule Interview", "✅ Update Result", "📋 All Interviews"],
        horizontal=True,
        key="interview_tab_selector"
    )
    
    st.markdown("---")
    
    # ========== DASHBOARD SECTION ==========
    if selected_tab == "📊 Dashboard":
        #st.markdown("### 📊 Interview Dashboard")
        
        if len(interviews_df) > 0:
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                matched_count = len(interviews_df[interviews_df['Interview Status'] == 'Matched'])
                st.metric("🎯 Matched", matched_count)
            
            with col2:
                scheduled_count = len(interviews_df[interviews_df['Interview Status'] == 'Interview Scheduled'])
                st.metric("🗓️ Scheduled", scheduled_count)
            
            with col3:
                completed_count = len(interviews_df[interviews_df['Interview Status'] == 'Interview Completed'])
                st.metric("✅ Completed", completed_count)
            
            with col4:
                selected_count = len(interviews_df[interviews_df['Result Status'] == 'Selected'])
                st.metric("🎉 Selected", selected_count)
            
            st.markdown("---")
            
            # Today's Interviews
            today = pd.Timestamp.now().strftime('%Y-%m-%d')
            
            if 'Interview Date' in interviews_df.columns:
                interviews_df['Interview Date'] = interviews_df['Interview Date'].astype(str)
                today_interviews = interviews_df[
                    (interviews_df['Interview Status'] == 'Interview Scheduled') & 
                    (interviews_df['Interview Date'].str.contains(today, na=False))
                ]
                
                if len(today_interviews) > 0:
                    st.markdown("### 🔥 Today's Interviews")
                    display_cols = ['Record ID', 'Full Name', 'Company Name', 'Job Title', 'Interview Time']
                    available_cols = [col for col in display_cols if col in today_interviews.columns]
                    st.dataframe(today_interviews[available_cols], use_container_width=True, hide_index=True)
                else:
                    st.info("✅ No interviews scheduled for today")
            else:
                st.info("✅ No interviews scheduled for today")
            
            st.markdown("---")
            
            # Pending Actions
            pending = interviews_df[interviews_df['Interview Status'] == 'Matched']
            if len(pending) > 0:
                st.markdown("### ⚠️ Pending to Schedule")
                st.warning(f"{len(pending)} interviews need to be scheduled!")
                display_cols = ['Record ID', 'Full Name', 'Company Name', 'Job Title', 'Match Score']
                available_cols = [col for col in display_cols if col in pending.columns]
                st.dataframe(pending[available_cols], use_container_width=True, hide_index=True)
        else:
            st.info("No interview records found. Export some matches from Job Matching to get started!")
    
    # ========== SCHEDULE INTERVIEW SECTION ==========
    elif selected_tab == "🗓️ Schedule Interview":
        st.markdown("### 🗓️ Schedule Interview")
        
        matched_interviews = get_schedulable_interviews(interviews_df, vacancies_df)
        
        if len(matched_interviews) > 0:
            st.success(f"📊 {len(matched_interviews)} interviews ready to schedule")
            
            record_options = matched_interviews.apply(
                lambda x: f"{x['Record ID']} | {x['Full Name']} → {x['Company Name']} ({x['Job Title']})", 
                axis=1
            ).tolist()
            
            selected_record = st.selectbox(
                "Select Interview Record",
                record_options,
                key="schedule_select_interview"
            )
            
            if selected_record:
                record_id = selected_record.split('|')[0].strip()
                interview_data = matched_interviews[matched_interviews['Record ID'] == record_id].iloc[0]
                
                st.markdown("---")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("#### 👤 Candidate Details")
                    st.write(f"**Name:** {interview_data['Full Name']}")
                    st.write(f"**ID:** {interview_data['Candidate ID']}")
                    
                    cand_id_col = 'Candidate_ID' if 'Candidate_ID' in candidates_df.columns else 'Candidate ID'
                    candidate_info = candidates_df[candidates_df[cand_id_col] == interview_data['Candidate ID']]
                    
                    if len(candidate_info) > 0:
                        candidate = candidate_info.iloc[0]
                        phone = candidate.get('Phone', candidate.get('Mobile', candidate.get('Contact Number', 'N/A')))
                        email = candidate.get('Email', 'N/A')
                        st.write(f"**📞 Phone:** {phone}")
                        st.write(f"**📧 Email:** {email}")
                    else:
                        st.warning("⚠️ Candidate details not found")
                
                with col2:
                    st.markdown("#### 🏢 Company Details")
                    st.write(f"**Company:** {interview_data['Company Name']}")
                    st.write(f"**Position:** {interview_data['Job Title']}")
                    st.write(f"**Match Score:** {interview_data['Match Score']}")
                    
                    company_info = companies_df[companies_df['CID'] == interview_data['CID']]
                    
                    if len(company_info) > 0:
                        company = company_info.iloc[0]
                        st.write(f"**👤 Contact:** {company.get('Contact Person', 'N/A')}")
                        st.write(f"**📞 Phone:** {company.get('Contact Number', 'N/A')}")
                        st.write(f"**📍 Address:** {company.get('Address of Company', company.get('Address', 'N/A'))}")
                    else:
                        st.warning("⚠️ Company details not found")
                
                st.markdown("---")
                
                with st.form("schedule_interview_form"):
                    st.markdown("#### 📅 Schedule Details")
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        interview_date = st.date_input(
                            "Interview Date *",
                            min_value=pd.Timestamp.now().date(),
                            value=pd.Timestamp.now().date() + pd.Timedelta(days=1)
                        )
                        interview_time = st.time_input("Interview Time *", value=pd.Timestamp('10:00').time())
                        round_number = st.selectbox("Interview Round *", ["Round 1", "Round 2", "Round 3", "Final Round"], index=0)
                    
                    with col2:
                        interview_mode = st.selectbox("Interview Mode *", ["Offline", "Online", "Hybrid"], index=0)
                        
                        if interview_mode == "Online":
                            meeting_link = st.text_input("Meeting Link (Google Meet/Zoom)")
                            interview_location = ""
                        else:
                            interview_location = st.text_input("Interview Location")
                            meeting_link = ""
                        
                        interviewer_name = st.text_input("Interviewer Name")
                    
                    remarks = st.text_area("Additional Remarks", height=100)
                    submit_schedule = st.form_submit_button("📅 Schedule Interview", type="primary")
                
                if submit_schedule:
                    try:
                        client = get_google_sheets_client()
                        if client:
                            SHEET_ID = get_agency_sheet_id()
                            if not SHEET_ID:
                                st.error("❌ Sheet not configured!")
                                return
                            
                            sheet = client.open_by_key(SHEET_ID).worksheet("Interview_Records")
                            all_data = sheet.get_all_values()
                            
                            if not all_data:
                                st.error("Interview_Records sheet is empty.")
                                return
                            
                            headers = all_data[0]
                            row_to_update = None
                            
                            for idx, row in enumerate(all_data[1:], start=2):
                                if row[0] == record_id:
                                    row_to_update = idx
                                    break
                            
                            if row_to_update:
                                updates = []
                                
                                status_col = headers.index('Interview Status') + 1 if 'Interview Status' in headers else 9
                                updates.append({'range': f"{chr(64 + status_col)}{row_to_update}", 'values': [['Interview Scheduled']]})
                                
                                date_col = headers.index('Interview Date') + 1 if 'Interview Date' in headers else 10
                                updates.append({'range': f"{chr(64 + date_col)}{row_to_update}", 'values': [[interview_date.strftime('%Y-%m-%d')]]})
                                
                                time_col = headers.index('Interview Time') + 1 if 'Interview Time' in headers else 11
                                updates.append({'range': f"{chr(64 + time_col)}{row_to_update}", 'values': [[interview_time.strftime('%H:%M')]]})
                                
                                round_col = headers.index('Interview Round') + 1 if 'Interview Round' in headers else 12
                                updates.append({'range': f"{chr(64 + round_col)}{row_to_update}", 'values': [[round_number]]})
                                
                                remarks_col = headers.index('Remarks') + 1 if 'Remarks' in headers else 16
                                location_info = meeting_link if interview_mode == "Online" else interview_location
                                full_remarks = f"Mode: {interview_mode} | Location/Link: {location_info} | Interviewer: {interviewer_name} | {remarks}"
                                updates.append({'range': f"{chr(64 + remarks_col)}{row_to_update}", 'values': [[full_remarks]]})
                                
                                updated_col = headers.index('Last Updated') + 1 if 'Last Updated' in headers else 17
                                updates.append({'range': f"{chr(64 + updated_col)}{row_to_update}", 'values': [[pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')]]})
                                
                                updated_by_col = headers.index('Updated By') + 1 if 'Updated By' in headers else 18
                                updates.append({'range': f"{chr(64 + updated_by_col)}{row_to_update}", 'values': [[st.session_state.get('username', 'Admin')]]})
                                
                                sheet.batch_update(updates)
                                
                                st.success("✅ Interview scheduled successfully!")
                                st.info("📧 Email notification will be sent automatically")
                                st.balloons()
                                st.cache_data.clear()
                                
                                # Keep the same tab selected after rerun
                                st.session_state.interview_tab_selector = "🗓️ Schedule Interview"
                                
                                import time
                                time.sleep(2)
                                st.rerun()
                            else:
                                st.error("❌ Record not found")
                        else:
                            st.error("❌ Could not connect to Google Sheets")
                    except Exception as e:
                        st.error(f"❌ Error: {str(e)}")
        else:
            st.info("✅ No interviews to schedule!")
    
    # ========== UPDATE RESULT SECTION ==========
    elif selected_tab == "✅ Update Result":
        st.markdown("### ✅ Update Interview Result")
        
        if len(vacancies_df) > 0 and 'CID' in vacancies_df.columns and 'Job Title' in vacancies_df.columns and 'status' in vacancies_df.columns:
            interviews_df = interviews_df.merge(vacancies_df[['CID', 'Job Title', 'status']], on=['CID', 'Job Title'], how='left')
        
        if len(interviews_df) > 0:
            updatable_interviews = get_updatable_interviews(interviews_df, vacancies_df)
            
            if len(updatable_interviews) > 0:
                st.info(f"📊 {len(updatable_interviews)} interviews to update")
                
                record_options = updatable_interviews.apply(
                    lambda x: f"{x['Record ID']} | {x['Full Name']} → {x['Company Name']} | {x.get('Interview Date', 'No Date')} {x.get('Interview Time', '')}", 
                    axis=1
                ).tolist()
                
                selected_record = st.selectbox("Select Interview Record", record_options, key="update_select_interview")
                
                if selected_record:
                    record_id = selected_record.split('|')[0].strip()
                    interview_data = updatable_interviews[updatable_interviews['Record ID'] == record_id].iloc[0]
                    
                    st.markdown("---")
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        st.write(f"**Candidate:** {interview_data['Full Name']}")
                        st.write(f"**Company:** {interview_data['Company Name']}")
                        st.write(f"**Position:** {interview_data['Job Title']}")
                    with col2:
                        st.write(f"**Date:** {interview_data.get('Interview Date', 'N/A')}")
                        st.write(f"**Time:** {interview_data.get('Interview Time', 'N/A')}")
                        st.write(f"**Round:** {interview_data.get('Interview Round', 'N/A')}")
                    
                    st.markdown("---")
                    
                    with st.form("update_result_form"):
                        st.markdown("#### 📝 Update Result")
                        
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            interview_status = st.selectbox("Interview Status *", 
                                ["Interview Scheduled", "Interview Completed", "Cancelled", "Rescheduled"], index=1)
                            result_status = st.selectbox("Result Status *", 
                                ["Pending", "Selected", "Rejected", "On Hold", "Next Round", "Candidate Declined", "Company Declined"], index=0)
                        
                        with col2:
                            if result_status == "Selected":
                                salary_offered = st.number_input("Salary Offered (₹)", min_value=0, step=1000,
                                    value=int(interview_data.get('Salary Offered', 0)) if pd.notna(interview_data.get('Salary Offered')) and interview_data.get('Salary Offered') != '' else 0)
                                joining_date = st.date_input("Joining Date", value=pd.Timestamp.now().date() + pd.Timedelta(days=15))
                            else:
                                salary_offered = None
                                joining_date = None
                        
                        feedback = st.text_area("Feedback/Remarks", height=150)
                        submit_result = st.form_submit_button("💾 Update Result", type="primary")
                    
                    if submit_result:
                        existing_selections = []
                        choice = 'proceed'
                        
                        if result_status == "Selected":
                            cancel_pending_entries(interview_data['Candidate ID'], record_id)
                            existing_selections = check_existing_selections(interview_data['Candidate ID'])
                            
                            if existing_selections:
                                st.warning("⚠️ This candidate already has selection(s)!")
                                st.write("### Existing Selection(s):")
                                for sel in existing_selections:
                                    st.write(f"- **{sel['company']}** | {sel['job_title']}")
                                
                                st.write("---")
                                st.write("### What do you want to do?")
                                
                                col1, col2 = st.columns(2)
                                with col1:
                                    if st.button("✅ Keep NEW Selection (Reject old)", key="keep_new_sel"):
                                        choice = 'current'
                                with col2:
                                    if st.button("✅ Keep OLD Selection (Reject new)", key="keep_old_sel"):
                                        choice = 'existing'
                                
                                if choice == 'proceed':
                                    st.info("👆 Select an option above to proceed")
                                    st.stop()
                        
                        if choice != 'proceed' or result_status != "Selected" or not existing_selections:
                            try:
                                client = get_google_sheets_client()
                                if client:
                                    SHEET_ID = get_agency_sheet_id()
                                    if not SHEET_ID:
                                        st.error("❌ Sheet not configured!")
                                        return
                                    
                                    sheet = client.open_by_key(SHEET_ID).worksheet("Interview_Records")
                                    all_data = sheet.get_all_values()
                                    
                                    if not all_data:
                                        st.error("Interview_Records sheet is empty.")
                                        return
                                    
                                    headers = all_data[0]
                                    row_to_update = None
                                    
                                    for idx, row in enumerate(all_data[1:], start=2):
                                        if row[0] == record_id:
                                            row_to_update = idx
                                            break
                                    
                                    if row_to_update:
                                        updates = []
                                        
                                        status_col = headers.index('Interview Status') + 1 if 'Interview Status' in headers else 9
                                        updates.append({'range': f"{chr(64 + status_col)}{row_to_update}", 'values': [[interview_status]]})
                                        
                                        result_col = headers.index('Result Status') + 1 if 'Result Status' in headers else 13
                                        updates.append({'range': f"{chr(64 + result_col)}{row_to_update}", 'values': [[result_status]]})
                                        
                                        if salary_offered is not None:
                                            salary_col = headers.index('Salary Offered') + 1 if 'Salary Offered' in headers else 14
                                            updates.append({'range': f"{chr(64 + salary_col)}{row_to_update}", 'values': [[salary_offered]]})
                                        
                                        if joining_date is not None:
                                            joining_col = headers.index('Joining Date') + 1 if 'Joining Date' in headers else 15
                                            updates.append({'range': f"{chr(64 + joining_col)}{row_to_update}", 'values': [[joining_date.strftime('%Y-%m-%d')]]})
                                        
                                        remarks_col = headers.index('Remarks') + 1 if 'Remarks' in headers else 16
                                        existing_remarks = interview_data.get('Remarks', '')
                                        new_remarks = f"{existing_remarks}\n\n[{pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}] {feedback}"
                                        updates.append({'range': f"{chr(64 + remarks_col)}{row_to_update}", 'values': [[new_remarks]]})
                                        
                                        updated_col = headers.index('Last Updated') + 1 if 'Last Updated' in headers else 17
                                        updates.append({'range': f"{chr(64 + updated_col)}{row_to_update}", 'values': [[pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')]]})
                                        
                                        updated_by_col = headers.index('Updated By') + 1 if 'Updated By' in headers else 18
                                        updates.append({'range': f"{chr(64 + updated_by_col)}{row_to_update}", 'values': [[st.session_state.get('username', 'Admin')]]})
                                        
                                        sheet.batch_update(updates)
                                        
                                        if result_status == "Selected" and existing_selections:
                                            update_selection_status(record_id, choice, existing_selections)
                                        
                                        st.success("✅ Result updated in Interview_Records!")
                                        
                                        sync_result = sync_all_statuses(
                                            candidate_id=interview_data['Candidate ID'],
                                            company_id=interview_data['CID'],
                                            job_title=interview_data['Job Title'],
                                            interview_status=interview_status,
                                            result_status=result_status
                                        )
                                        
                                        if sync_result:
                                            st.success("✅ Candidate status and vacancy status synced!")
                                        else:
                                            st.warning("⚠️ Result updated but some sync issues occurred")
                                        
                                        if result_status == "Selected":
                                            st.balloons()
                                        
                                        st.cache_data.clear()
                                        
                                        # Keep the same tab selected after rerun
                                        st.session_state.interview_tab_selector = "✅ Update Result"
                                        
                                        import time
                                        time.sleep(2)
                                        st.rerun()
                                    else:
                                        st.error("❌ Record not found")
                                else:
                                    st.error("❌ Could not connect to Google Sheets")
                            except Exception as e:
                                st.error(f"❌ Error: {str(e)}")
            else:
                st.info("✅ No interviews to update")
        else:
            st.info("No interview records found")
    
    # ========== ALL INTERVIEWS SECTION ==========
    elif selected_tab == "📋 All Interviews":
        st.markdown("### 📋 All Interview Records")
        
        if len(interviews_df) > 0:
            col1, col2, col3 = st.columns(3)
            
            with col1:
                if 'Interview Status' in interviews_df.columns:
                    status_filter = st.multiselect("Filter by Interview Status",
                        options=interviews_df['Interview Status'].unique().tolist(),
                        default=interviews_df['Interview Status'].unique().tolist())
                else:
                    status_filter = []
            
            with col2:
                if 'Result Status' in interviews_df.columns:
                    result_filter = st.multiselect("Filter by Result Status",
                        options=interviews_df['Result Status'].unique().tolist(),
                        default=interviews_df['Result Status'].unique().tolist())
                else:
                    result_filter = []
            
            with col3:
                search_text = st.text_input("Search (Name/Company)")
            
            filtered_df = interviews_df.copy()
            
            if status_filter and 'Interview Status' in interviews_df.columns:
                filtered_df = filtered_df[filtered_df['Interview Status'].isin(status_filter)]
            
            if result_filter and 'Result Status' in interviews_df.columns:
                filtered_df = filtered_df[filtered_df['Result Status'].isin(result_filter)]
            
            if search_text:
                if 'Full Name' in filtered_df.columns and 'Company Name' in filtered_df.columns:
                    filtered_df = filtered_df[
                        filtered_df['Full Name'].str.contains(search_text, case=False, na=False) |
                        filtered_df['Company Name'].str.contains(search_text, case=False, na=False)
                    ]
            
            st.write(f"**Showing {len(filtered_df)} / {len(interviews_df)} records**")
            st.dataframe(filtered_df, use_container_width=True, height=400)
            
            csv = filtered_df.to_csv(index=False)
            st.download_button(
                label="📥 Download Filtered Data (CSV)",
                data=csv,
                file_name=f"interviews_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv"
            )
        else:
            st.info("No interview records found. Export some matches from Job Matching to get started!")
# ==================================================== 
# REPORTS
# ==================================================== 
def admin_reports():
    render_page_header("Reports & Analytics", "View performance metrics and generate reports")
    tab1, tab2, tab3 = st.tabs(["Today", "Period", "Overall"])
    
    with tab1:
        st.markdown("### 📅 Today's Activity")
        today = pd.Timestamp.now().date()
        
        candidates_df = get_candidates()
        interviews_df = get_interviews()
        vacancies_df = get_vacancies()
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            if len(candidates_df) > 0 and 'Date Applied' in candidates_df.columns:
                candidates_df['Date Applied'] = pd.to_datetime(candidates_df['Date Applied'], errors='coerce')
                today_candidates = len(candidates_df[candidates_df['Date Applied'].dt.date == today])
            else:
                today_candidates = 0
            st.metric("👥 New Candidates", today_candidates, delta="Today")
        
        with col2:
            if len(interviews_df) > 0 and 'Interview Date' in interviews_df.columns:
                interviews_df['Interview Date'] = pd.to_datetime(interviews_df['Interview Date'], errors='coerce')
                today_interviews = len(interviews_df[(interviews_df['Interview Date'].dt.date == today) & 
                                                     (interviews_df['Interview Status'] == 'Interview Scheduled')])
            else:
                today_interviews = 0
            st.metric("🗓️ Interviews", today_interviews, delta="Scheduled")
        
        with col3:
            if len(interviews_df) > 0 and 'Last Updated' in interviews_df.columns:
                interviews_df['Last Updated'] = pd.to_datetime(interviews_df['Last Updated'], errors='coerce')
                today_selected = len(interviews_df[(interviews_df['Last Updated'].dt.date == today) & 
                                                   (interviews_df['Result Status'] == 'Selected')])
            else:
                today_selected = 0
            st.metric("🎉 Selected", today_selected, delta="Placements")
        
        with col4:
            if len(vacancies_df) > 0 and 'Date Added' in vacancies_df.columns:
                vacancies_df['Date Added'] = pd.to_datetime(vacancies_df['Date Added'], errors='coerce')
                today_vacancies = len(vacancies_df[vacancies_df['Date Added'].dt.date == today])
            else:
                today_vacancies = 0
            st.metric("💼 New Vacancies", today_vacancies, delta="Posted")
    
    with tab2:
        st.markdown("### 📊 Period Summary")
        st.info("Period reports - Full implementation available")
    
    with tab3:
        st.markdown("### 📈 Overall Statistics")
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Companies", len(get_companies()))
        col2.metric("Vacancies", len(get_vacancies()))
        col3.metric("Candidates", len(get_candidates()))
        col4.metric("Interviews", len(get_interviews()))

def extract_sheet_id(url):
    import re
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
    return match.group(1) if match else None


def get_next_record_id(spreadsheet, sheet_name, prefix):
    try:
        sheet = spreadsheet.worksheet(sheet_name)
        records = sheet.get_all_records()
        if not records:
            return f"{prefix}001"
        first_col = list(records[0].keys())[0]
        ids = [str(r.get(first_col, "")) for r in records if str(r.get(first_col, "")).startswith(prefix)]
        if not ids:
            return f"{prefix}001"
        nums = []
        for i in ids:
            try:
                nums.append(int(i.replace(prefix, "")))
            except Exception:
                pass
        next_num = max(nums) + 1 if nums else 1
        return f"{prefix}{next_num:03d}"
    except Exception:
        return f"{prefix}001"


def get_company_list(spreadsheet):
    try:
        sheet = spreadsheet.worksheet("CID")
        records = sheet.get_all_records()
        df = pd.DataFrame(records)
        if df.empty:
            return pd.DataFrame(columns=["CID", "Company_Name", "Subscription_Status",
                                          "Subscription_End_Date", "Plan_Type"])
        for col in ["Subscription_Status", "Subscription_End_Date", "Plan_Type"]:
            if col not in df.columns:
                df[col] = ""
        return df
    except Exception as e:
        st.error(f"Error reading CID sheet: {e}")
        return pd.DataFrame()


def get_active_plans(spreadsheet):
    try:
        sheet = spreadsheet.worksheet("Subscription_Plans")
        records = sheet.get_all_records()
        df = pd.DataFrame(records)
        if df.empty:
            return pd.DataFrame()
        df = df[df["Is_Active"].astype(str).str.upper() == "YES"]
        return df.reset_index(drop=True)
    except Exception as e:
        st.error(f"Error reading Subscription_Plans sheet: {e}")
        return pd.DataFrame()


def _ensure_cid_subscription_columns(sheet):
    headers = sheet.row_values(1)
    needed = ["Subscription_Status", "Subscription_End_Date", "Plan_Type"]
    for col_name in needed:
        if col_name not in headers:
            headers.append(col_name)
            col_letter = get_column_letter(len(headers))
            sheet.update(f"{col_letter}1", col_name)
    return headers


def update_cid_subscription(spreadsheet, company_cid, status, end_date, plan_type):
    try:
        sheet = spreadsheet.worksheet("CID")
        headers = _ensure_cid_subscription_columns(sheet)
        all_values = sheet.get_all_values()
        if not all_values or len(all_values) < 2:
            return False, "CID sheet has no data."
        header_row = all_values[0]
        try:
            cid_col_idx = header_row.index("CID") + 1
            status_col_idx = header_row.index("Subscription_Status") + 1
            end_date_col_idx = header_row.index("Subscription_End_Date") + 1
            plan_col_idx = header_row.index("Plan_Type") + 1
        except ValueError as e:
            return False, f"Missing column in CID sheet: {e}"
        target_row = None
        for i, row in enumerate(all_values[1:], start=2):
            if len(row) >= cid_col_idx and str(row[cid_col_idx - 1]).strip() == str(company_cid).strip():
                target_row = i
                break
        if not target_row:
            return False, f"Company CID '{company_cid}' not found in CID sheet."
        end_date_str = end_date.strftime("%Y-%m-%d") if hasattr(end_date, "strftime") else str(end_date)
        sheet.update_cell(target_row, status_col_idx, status)
        sheet.update_cell(target_row, end_date_col_idx, end_date_str)
        sheet.update_cell(target_row, plan_col_idx, plan_type)
        return True, "CID sheet updated successfully."
    except Exception as e:
        return False, f"Error updating CID sheet: {e}"


def expire_old_subscription(spreadsheet, company_cid):
    try:
        sheet = spreadsheet.worksheet("Subscription_Records")
        all_values = sheet.get_all_values()
        if len(all_values) < 2:
            return 0
        headers = all_values[0]
        try:
            cid_col = headers.index("Company_CID") + 1
            status_col = headers.index("Status") + 1
            updated_col = headers.index("Last_Updated") + 1
        except ValueError:
            return 0
        count = 0
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for i, row in enumerate(all_values[1:], start=2):
            if (len(row) >= max(cid_col, status_col) and
                    str(row[cid_col - 1]).strip() == str(company_cid).strip() and
                    str(row[status_col - 1]).strip().upper() == "ACTIVE"):
                sheet.update_cell(i, status_col, "EXPIRED")
                sheet.update_cell(i, updated_col, now_str)
                count += 1
        return count
    except Exception as e:
        st.warning(f"Could not expire old subscription: {e}")
        return 0


def add_subscription_record(spreadsheet, company_cid, company_name, plan_type,
                             duration_months, amount, payment_date, payment_mode,
                             transaction_id, invoice_no, notes, created_by):
    try:
        sub_sheet = spreadsheet.worksheet("Subscription_Records")
        records = sub_sheet.get_all_records()
        start_date = date.today()
        for r in records:
            if (str(r.get("Company_CID", "")).strip() == str(company_cid).strip() and
                    str(r.get("Status", "")).strip().upper() == "ACTIVE"):
                try:
                    existing_end = datetime.strptime(str(r["End_Date"]).strip(), "%Y-%m-%d").date()
                    if existing_end > start_date:
                        start_date = existing_end
                except Exception:
                    pass
        month = start_date.month - 1 + int(duration_months)
        year = start_date.year + month // 12
        month = month % 12 + 1
        try:
            end_date = start_date.replace(year=year, month=month)
        except ValueError:
            import calendar
            last_day = calendar.monthrange(year, month)[1]
            end_date = start_date.replace(year=year, month=month, day=last_day)
        expire_old_subscription(spreadsheet, company_cid)
        record_id = get_next_record_id(spreadsheet, "Subscription_Records", "SUB")
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        payment_date_str = payment_date.strftime("%Y-%m-%d") if hasattr(payment_date, "strftime") else str(payment_date)
        new_row = [
            record_id, company_cid, company_name, plan_type, amount,
            start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d"),
            "ACTIVE", payment_date_str, payment_mode, transaction_id,
            invoice_no, notes, created_by, now_str, now_str, created_by,
        ]
        sub_sheet.append_row(new_row)
        ok, msg = update_cid_subscription(spreadsheet, company_cid, "ACTIVE", end_date, plan_type)
        if not ok:
            return True, f"Subscription added but CID update failed: {msg}", record_id
        return True, "Subscription added successfully.", record_id
    except Exception as e:
        return False, f"Error adding subscription: {e}", ""


def check_and_expire_subscriptions(spreadsheet):
    expired_count = 0
    messages = []
    today = date.today()
    try:
        sheet = spreadsheet.worksheet("Subscription_Records")
        all_values = sheet.get_all_values()
        if len(all_values) < 2:
            return 0, ["No subscription records found."]
        headers = all_values[0]
        try:
            status_col = headers.index("Status") + 1
            end_date_col = headers.index("End_Date") + 1
            cid_col = headers.index("Company_CID") + 1
            plan_col = headers.index("Plan_Type") + 1
            updated_col = headers.index("Last_Updated") + 1
        except ValueError as e:
            return 0, [f"Missing column: {e}"]
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for i, row in enumerate(all_values[1:], start=2):
            if len(row) < max(status_col, end_date_col):
                continue
            if str(row[status_col - 1]).strip().upper() != "ACTIVE":
                continue
            try:
                end_date = datetime.strptime(str(row[end_date_col - 1]).strip(), "%Y-%m-%d").date()
            except Exception:
                continue
            if end_date < today:
                sheet.update_cell(i, status_col, "EXPIRED")
                sheet.update_cell(i, updated_col, now_str)
                company_cid = row[cid_col - 1]
                plan_type = row[plan_col - 1] if len(row) >= plan_col else ""
                update_cid_subscription(spreadsheet, company_cid, "EXPIRED", end_date, plan_type)
                expired_count += 1
                messages.append(f"Expired: {company_cid} (End: {end_date})")
        if expired_count == 0:
            messages.append("No subscriptions needed expiration.")
        return expired_count, messages
    except Exception as e:
        return 0, [f"Error: {e}"]   

def render_settings():
    st.markdown("### ⚙️ Agency Settings")

    agency_sheet_url = st.session_state.get("agency_sheet_url", "")
    if not agency_sheet_url:
        st.error("No agency sheet URL found in session. Please log in again.")
        return

    sheet_id = extract_sheet_id(agency_sheet_url)
    if not sheet_id:
        st.error("Invalid Google Sheet URL in session.")
        return

    try:
        client = get_google_sheets_client()
        spreadsheet = client.open_by_key(sheet_id)
    except Exception as e:
        st.error(f"Could not open agency spreadsheet: {e}")
        return

    current_user = st.session_state.get("username", "Admin")
    agency_code = st.session_state.get("agency_code", "")

    tab1, tab2, tab3, tab4 = st.tabs([
        "💰 Manage Subscriptions",
        "💳 Pricing Plans",
        "⏰ Background Tasks",
        "🔐 Agency Profile",
    ])

    # ── TAB 1 ──────────────────────────────────────────────────────
    with tab1:
        st.subheader("Active Subscriptions")

        try:
            sub_sheet = spreadsheet.worksheet("Subscription_Records")
            sub_records = sub_sheet.get_all_records()
        except Exception as e:
            st.warning(f"Could not read Subscription_Records: {e}")
            sub_records = []

        today = date.today()

        active_subs = [r for r in sub_records if str(r.get("Status", "")).upper() == "ACTIVE"]
        if active_subs:
            rows = []
            for r in active_subs:
                try:
                    end_d = datetime.strptime(str(r["End_Date"]).strip(), "%Y-%m-%d").date()
                    days_left = (end_d - today).days
                except Exception:
                    days_left = "N/A"
                rows.append({
                    "CID": r.get("Company_CID", ""),
                    "Company": r.get("Company_Name", ""),
                    "Plan": r.get("Plan_Type", ""),
                    "End Date": r.get("End_Date", ""),
                    "Days Left": days_left,
                    "Status": r.get("Status", ""),
                })
            df_active = pd.DataFrame(rows)
            st.dataframe(df_active, use_container_width=True, hide_index=True)
        else:
            st.info("No active subscriptions found.")

        st.divider()
        st.subheader("Add / Renew Subscription")

        company_df = get_company_list(spreadsheet)
        if company_df.empty:
            st.warning("No companies found in CID sheet.")
        else:
            with st.expander("📋 All Companies & Subscription Status", expanded=False):
                display_cols = [c for c in ["CID", "Company_Name", "Subscription_Status",
                                             "Subscription_End_Date", "Plan_Type"]
                                if c in company_df.columns]
                st.dataframe(company_df[display_cols], use_container_width=True, hide_index=True)

            company_options = ["-- Select Company --"] + [
                f"{row.get('Company_Name', '')} ({row.get('CID', '')})"
                for _, row in company_df.iterrows()
            ]

            plans_df = get_active_plans(spreadsheet)
            if plans_df.empty:
                st.warning("No active pricing plans found in Subscription_Plans sheet.")
                plan_options = []
            else:
                plan_options = [
                    f"{row['Plan_Name']} – {row['Duration_Months']} months – ₹{row['Price']}"
                    for _, row in plans_df.iterrows()
                ]

            next_invoice = get_next_record_id(spreadsheet, "Subscription_Records", "INV")

            with st.form("add_subscription_form", clear_on_submit=True):
                st.markdown("#### Subscription Details")

                col1, col2 = st.columns(2)
                with col1:
                    selected_company_str = st.selectbox("Company *", company_options)
                with col2:
                    if plan_options:
                        selected_plan_str = st.selectbox("Plan *", ["-- Select Plan --"] + plan_options)
                    else:
                        selected_plan_str = st.selectbox("Plan *", ["No plans available"])

                col3, col4 = st.columns(2)
                with col3:
                    default_amount = 0
                    if plan_options and selected_plan_str not in ("-- Select Plan --", "No plans available"):
                        if selected_plan_str in plan_options:
                            plan_idx = plan_options.index(selected_plan_str)
                            default_amount = int(plans_df.iloc[plan_idx]["Price"])
                    amount = st.number_input("Amount Paid (₹) *", min_value=0, value=default_amount, step=100)
                with col4:
                    payment_date = st.date_input("Payment Date *", value=today)

                col5, col6 = st.columns(2)
                with col5:
                    payment_mode = st.selectbox("Payment Mode *", ["Cash", "Bank Transfer", "UPI", "Cheque"])
                with col6:
                    transaction_id = st.text_input("Transaction ID (optional)")

                col7, _ = st.columns(2)
                with col7:
                    invoice_no = st.text_input("Invoice No", value=next_invoice)

                notes = st.text_area("Notes (optional)", height=80)
                submitted = st.form_submit_button("✅ Add / Renew Subscription", use_container_width=True)

                if submitted:
                    errors = []
                    if selected_company_str == "-- Select Company --":
                        errors.append("Please select a company.")
                    if not plan_options or selected_plan_str in ("-- Select Plan --", "No plans available"):
                        errors.append("Please select a valid plan.")

                    if errors:
                        for e in errors:
                            st.error(e)
                    else:
                        import re as _re
                        cid_match = _re.search(r"\(([^)]+)\)$", selected_company_str)
                        company_cid = cid_match.group(1) if cid_match else ""
                        company_name = selected_company_str.replace(f" ({company_cid})", "").strip()

                        plan_idx = plan_options.index(selected_plan_str)
                        plan_row = plans_df.iloc[plan_idx]
                        plan_type = str(plan_row["Plan_Name"])
                        duration_months = int(plan_row["Duration_Months"])

                        ok, msg, record_id = add_subscription_record(
                            spreadsheet=spreadsheet,
                            company_cid=company_cid,
                            company_name=company_name,
                            plan_type=plan_type,
                            duration_months=duration_months,
                            amount=amount,
                            payment_date=payment_date,
                            payment_mode=payment_mode,
                            transaction_id=transaction_id,
                            invoice_no=invoice_no,
                            notes=notes,
                            created_by=current_user,
                        )
                        if ok:
                            st.success(f"✅ {msg} Record ID: {record_id}")
                            st.rerun()
                        else:
                            st.error(f"❌ {msg}")

    # ── TAB 2 ──────────────────────────────────────────────────────
    with tab2:
        st.subheader("Current Pricing Plans")

        plans_df = get_active_plans(spreadsheet)

        if not plans_df.empty:
            # String columns force karo
            display_plans = plans_df[["Plan_ID", "Plan_Name", "Duration_Months", "Price",
                        "Max_Vacancies", "Max_Candidate_Views", "Max_Contact_Requests", "Features"]].copy()

            for col in display_plans.columns:
                display_plans[col] = display_plans[col].astype(str)

            st.dataframe(
                display_plans,
                use_container_width=True,
                hide_index=True,
            )

            st.divider()
            st.subheader("✏️ Edit Existing Plan")
            plan_names = plans_df["Plan_Name"].tolist()
            edit_plan_name = st.selectbox("Select Plan to Edit", plan_names, key="edit_plan_select")
            edit_idx = plan_names.index(edit_plan_name)
            edit_row = plans_df.iloc[edit_idx]

            with st.form("edit_plan_form"):
                ec1, ec2 = st.columns(2)
                with ec1:
                    new_price = st.number_input("Price (₹)", min_value=0, value=int(edit_row.get("Price", 0)))
                    new_max_vac = st.text_input("Max Vacancies", value=str(edit_row.get("Max_Vacancies", "")))
                    new_max_views = st.text_input("Max Candidate Views", value=str(edit_row.get("Max_Candidate_Views", "")))
                with ec2:
                    new_max_contacts = st.text_input("Max Contact Requests", value=str(edit_row.get("Max_Contact_Requests", "")))
                    new_features = st.text_area("Features", value=str(edit_row.get("Features", "")), height=100)

                save_edit = st.form_submit_button("💾 Save Changes", use_container_width=True)
                if save_edit:
                    try:
                        plan_sheet = spreadsheet.worksheet("Subscription_Plans")
                        all_plan_values = plan_sheet.get_all_values()
                        plan_headers = all_plan_values[0]
                        plan_id_to_find = str(edit_row["Plan_ID"])
                        pid_col = plan_headers.index("Plan_ID") + 1
                        target_row = None
                        for i, row in enumerate(all_plan_values[1:], start=2):
                            if row[pid_col - 1] == plan_id_to_find:
                                target_row = i
                                break
                        if target_row:
                            col_map = {
                                "Price": new_price,
                                "Max_Vacancies": new_max_vac,
                                "Max_Candidate_Views": new_max_views,
                                "Max_Contact_Requests": new_max_contacts,
                                "Features": new_features,
                                "Last_Updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            }
                            for col_name, val in col_map.items():
                                if col_name in plan_headers:
                                    ci = plan_headers.index(col_name) + 1
                                    plan_sheet.update_cell(target_row, ci, val)
                            st.success("✅ Plan updated successfully.")
                            st.rerun()
                        else:
                            st.error("Plan not found in sheet.")
                    except Exception as e:
                        st.error(f"Error updating plan: {e}")
        else:
            st.info("No active plans found. Add a plan below.")

        st.divider()
        st.subheader("➕ Add Custom Plan")

        with st.form("add_plan_form", clear_on_submit=True):
            ac1, ac2 = st.columns(2)
            with ac1:
                new_plan_name = st.text_input("Plan Name *")
                new_duration = st.number_input("Duration (Months) *", min_value=1, max_value=60, value=3)
                new_plan_price = st.number_input("Price (₹) *", min_value=0, value=0)
            with ac2:
                new_plan_vac = st.text_input("Max Vacancies", value="50")
                new_plan_views = st.text_input("Max Candidate Views", value="500")
                new_plan_contacts = st.text_input("Max Contact Requests", value="100")
            new_plan_features = st.text_area("Features Description", height=80)

            add_plan_submit = st.form_submit_button("➕ Add Plan", use_container_width=True)
            if add_plan_submit:
                if not new_plan_name.strip():
                    st.error("Plan Name is required.")
                else:
                    try:
                        plan_sheet = spreadsheet.worksheet("Subscription_Plans")
                        new_plan_id = get_next_record_id(spreadsheet, "Subscription_Plans", "PLAN")
                        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        plan_sheet.append_row([
                            new_plan_id, new_plan_name.strip(), new_duration, new_plan_price,
                            new_plan_vac, new_plan_views, new_plan_contacts,
                            new_plan_features, "YES", now_str, now_str
                        ])
                        st.success(f"✅ Plan '{new_plan_name}' added with ID {new_plan_id}.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error adding plan: {e}")

    # ── TAB 3 ──────────────────────────────────────────────────────
    with tab3:
        st.subheader("⏰ Auto-Expire Subscriptions")
        st.markdown(
            "यह सभी **ACTIVE** subscriptions check करता है जिनकी **End Date** निकल गई हो "
            "और उन्हें **EXPIRED** mark करता है।"
        )

        if "last_expire_run" not in st.session_state:
            st.session_state["last_expire_run"] = None
        if "last_expire_results" not in st.session_state:
            st.session_state["last_expire_results"] = []
        if "last_expire_count" not in st.session_state:
            st.session_state["last_expire_count"] = 0

        last_run = st.session_state["last_expire_run"]
        if last_run:
            st.info(f"Last run: **{last_run}**  |  Expired: **{st.session_state['last_expire_count']}** subscription(s)")

        if st.button("▶️ Run Expiry Check Now", use_container_width=True):
            with st.spinner("Checking subscriptions..."):
                count, messages = check_and_expire_subscriptions(spreadsheet)
            st.session_state["last_expire_run"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            st.session_state["last_expire_count"] = count
            st.session_state["last_expire_results"] = messages
            if count > 0:
                st.success(f"✅ {count} subscription(s) marked as EXPIRED.")
            else:
                st.success("✅ Check complete. No subscriptions required expiration.")
            st.rerun()

        if st.session_state["last_expire_results"]:
            with st.expander("📋 Detailed Results", expanded=False):
                for msg in st.session_state["last_expire_results"]:
                    st.text(msg)

    # ── TAB 4 ──────────────────────────────────────────────────────
    with tab4:
        st.subheader("🔐 Agency Profile")

        agency_name = ""
        agency_email = ""
        agency_logo = ""
        agency_row_idx = None
        login_sheet = None
        login_headers = []

        try:
            login_sheet = spreadsheet.worksheet("login_master")
            all_login = login_sheet.get_all_values()
            if len(all_login) > 1:
                login_headers = all_login[0]
                for i, row in enumerate(all_login[1:], start=2):
                    row_dict = dict(zip(login_headers, row))
                    if str(row_dict.get("agency_code", "")).strip() == str(agency_code).strip():
                        agency_name = row_dict.get("agency_name", "")
                        agency_email = row_dict.get("email", "")
                        agency_logo = row_dict.get("logo_url", "")
                        agency_row_idx = i
                        break
        except Exception:
            st.info("Could not read login_master sheet. Profile editing unavailable.")

        if agency_logo:
            st.image(agency_logo, width=120)

        with st.form("agency_profile_form"):
            p1, p2 = st.columns(2)
            with p1:
                upd_name = st.text_input("Agency Name", value=agency_name)
                upd_email = st.text_input("Email", value=agency_email)
            with p2:
                upd_logo = st.text_input("Logo URL", value=agency_logo)

            save_profile = st.form_submit_button("💾 Save Profile", use_container_width=True)
            if save_profile:
                if login_sheet and agency_row_idx and login_headers:
                    try:
                        for field, val in [("agency_name", upd_name), ("email", upd_email), ("logo_url", upd_logo)]:
                            if field in login_headers:
                                ci = login_headers.index(field) + 1
                                login_sheet.update_cell(agency_row_idx, ci, val)
                        st.success("✅ Profile updated successfully.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error saving profile: {e}")
                else:
                    st.warning("Agency record not found. Profile could not be saved.")

        st.divider()
        st.subheader("🔑 Change Password")

        with st.form("change_password_form"):
            old_pw = st.text_input("Current Password", type="password")
            new_pw = st.text_input("New Password", type="password")
            confirm_pw = st.text_input("Confirm New Password", type="password")

            change_pw_submit = st.form_submit_button("🔒 Change Password", use_container_width=True)
            if change_pw_submit:
                if not old_pw or not new_pw or not confirm_pw:
                    st.error("All password fields are required.")
                elif new_pw != confirm_pw:
                    st.error("New passwords do not match.")
                elif len(new_pw) < 6:
                    st.error("Password must be at least 6 characters.")
                else:
                    if login_sheet and agency_row_idx and login_headers:
                        try:
                            pw_col = "password" if "password" in login_headers else "Password"
                            if pw_col in login_headers:
                                ci = login_headers.index(pw_col) + 1
                                current_stored_pw = login_sheet.cell(agency_row_idx, ci).value
                                if str(current_stored_pw).strip() != str(old_pw).strip():
                                    st.error("Current password is incorrect.")
                                else:
                                    login_sheet.update_cell(agency_row_idx, ci, new_pw)
                                    st.success("✅ Password changed successfully.")
                            else:
                                st.error("Password column not found in login_master.")
                        except Exception as e:
                            st.error(f"Error changing password: {e}")
                    else:
                        st.warning("Agency record not found. Password could not be changed.")



# ==================================================== 
# MAIN
# ==================================================== 
def main():
    if not st.session_state.get("logged_in", False):
        from login_master_with_branding import main as login_main
        login_main()
        return
        
    user_role = st.session_state.get("role", "").lower()
    full_name = st.session_state.get("full_name", "User")
    
    # ========== COMPANY USERS - SEPARATE FLOW ==========
    if user_role == "company":
        from company_portal import render_company_dashboard
        render_company_dashboard()
        return  # ✅ Exit - company has its own sidebar
    
    if user_role == "candidate":
        from candidate_portal import render_candidate_dashboard
        render_candidate_dashboard()
        return
        
    # ========== ADMIN/SUPER_ADMIN - ORIGINAL FLOW ==========
    menu_choice, logout_pressed, sidebar_role = render_sidebar()
    
    if logout_pressed:
        logout()
        return
    
    # SUPER ADMIN
    if user_role == "super_admin":
        if menu_choice == "Dashboard":
            render_page_header("Super Admin Dashboard", f"Welcome back, {full_name}")
        elif menu_choice == "Agency Management":
            st.info("Agency Management feature under development")
    
    # AGENCY ADMIN
    elif user_role == "admin":
        if menu_choice == "Dashboard":
            admin_dashboard()
        elif menu_choice == "Company Management":
            admin_company_mgmt()
        elif menu_choice == "Vacancy Management":
            admin_vacancy_mgmt()
        elif menu_choice == "Candidate Management":
            admin_candidate_mgmt()
        elif menu_choice == "Advanced Filtering":
            admin_advanced_filtering()
        elif menu_choice == "Job Matching":
            admin_job_matching()
        elif menu_choice == "Interview Management":
            admin_interview_mgmt()
        elif menu_choice == "Reports & Analytics":
            admin_reports()
        elif menu_choice == "Settings":
            #st.write(dict(st.session_state))  # debug line
            render_settings()    
    
    else:
        st.error("❌ Unknown role")

if __name__ == "__main__":
    main()